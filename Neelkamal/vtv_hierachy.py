import openpyxl
import psycopg2
from openpyxl import Workbook, load_workbook

#  Main Sheet
wb = Workbook()
ws = wb.create_chartsheet
ws = openpyxl.load_workbook(r"F:\sequelstring\Excel\NEELMETAL_OUTPUT\Output\vtv_hierachy.xlsx", data_only=True)
ws = wb['Sheet']

# VTV- 01102021
wb1 = openpyxl.load_workbook(r"F:\sequelstring\Excel\NEELMETAL_OUTPUT\PIR_FRAME_NCR_NEELMETAL_01.01.2022.xlsx", data_only=True)
vtv_sheet = wb1['VTV- 01102021']

# SUMMARY-FRAMEBODY
wb2 = openpyxl.load_workbook(r"F:\sequelstring\Excel\NEELMETAL_OUTPUT\PIR_FRAME_NCR_NEELMETAL_01.01.2022.xlsx", data_only=True)
summary_t = wb2['SUMMARY-FRAMEBODY']

wb3 = openpyxl.load_workbook(r"F:\sequelstring\Excel\NEELMETAL_OUTPUT\PIR_FRAME_NCR_NEELMETAL_01.01.2022.xlsx", data_only=False)
summary_f = wb3['SUMMARY-FRAMEBODY']

# NEELMETAL
wb4 = openpyxl.load_workbook(r"F:\sequelstring\Excel\NEELMETAL_OUTPUT\PIR_FRAME_NCR_NEELMETAL_01.01.2022.xlsx", data_only=True)
neel = wb4['NEELMETAL']

wb5 = openpyxl.load_workbook(r"F:\sequelstring\Excel\NEELMETAL_OUTPUT\PIR_FRAME_NCR_NEELMETAL_01.01.2022.xlsx", data_only=False)
neel_f = wb5['NEELMETAL']


headers = ['bom_hierarchy', 'direct_sup', 'purchase_group', 'plan',	'from_date', 'to_date',	'partcode', 'partcode_type1_oe',
'partcode_level2','partcode_type2_rm_bop_vtv_inm',	'partcode_level3', 'partcode_type3_rm_bop_vtv_inm',	'partcode_level4', 'partcode_type4_rm_bop_vtv_inm',	'partcode_level5','partcode_type5_rm_bop_vtv_inm','partcode_level6','partcode_type6_rm_bop_vtv_inm_base_reference_pc']
ws.append(headers)

# VTV- 01102021
for i in range(5, 38):
    vtv_part = vtv_sheet["G" + str(i)].value
    # print(vtv_part)
    revised_price = vtv_sheet["W" + str(i)].value
    direct_vendor = vtv_sheet["B" + str(i)].value
    plan = vtv_sheet["D" + str(i)].value
    from_date = "01.01.22"
    to_date = "31.12.9999"
    # print(vtv_part, i)
    # print(revised_price)

    # for summary framebody
    for j in range(9, 73):
        summary_part = summary_t["G" + str(j)].value
        # print("summary_part", summary_part)
        summary_price = summary_t["BD" + str(j)].value
        summary_price_f = summary_f["BD" + str(j)].value

        if vtv_part == summary_part and revised_price == summary_price:
            summary_interest_cost = summary_t["Z" + str(j)].value
            summary_depriciation_cost = summary_t["AA" + str(j)].value
            summary_Wire_and_Co2_gas = summary_t["AB" + str(j)].value
            summary_other_cost = summary_t["AC" + str(j)].value
            summary_welding_fix_cost = summary_t["AD" + str(j)].value
            summary_press_shop_cost = summary_t["AE" + str(j)].value
            summary_OH_at_30_cost = summary_t["AF" + str(j)].value
            summary_profit_cost = summary_t["AG" + str(j)].value
            summary_rejection_at_rm_and_conv_cost = summary_t["AH" + str(j)].value
            summary_paking_exp_cost = summary_t["AI" + str(j)].value
            summary_freight_cost = summary_t["AJ" + str(j)].value
            summary_ammortization_cost = summary_t["AK" + str(j)].value
            summary_power_welding = summary_t["V" + str(j)].value
            summary_power_press_shop = summary_t["W" + str(j)].value
            summary_labour_welding = summary_t["R" + str(j)].value
            summary_labour_press_shop = summary_t["S" + str(j)].value
            summary_labour_other_assy_operation = summary_t["T" + str(j)].value
            summary_rm_sec_operaton_charges = summary_t["AM" + str(20)].value
            summary_additional_price = summary_t["BL" + str(j)].value

            summary_rev_BOP = summary_t["K" + str(j)].value

            l1 = [ "", direct_vendor,"", plan, from_date, to_date, vtv_part, "OE"]
            ws.append(l1)


            # NEELMETAL
            for n in range(31, 4042):
                neel_part = neel["C" + str(n)].value
                neel_price = neel["AI" + str(n)].value

                if neel_part == summary_part and summary_rev_BOP == neel_price:
                    neel_part_f = neel_f["C" + str(n)].value
                    if neel_part_f == None:
                        continue
                    # print("neel_part_f", neel_part_f)
                    neel_price_f = neel_f["AI" + str(n)].value
                    if neel_price_f == None:
                        continue
                    # print("neel_price_f", neel_price_f)

                    neel_price_f = neel_price_f.replace("AI", "").replace("=", "")
                    neel_price_f = neel_f["AI" + str(neel_price_f)].value
                    # print(neel_price_f)
                    neel_price_f1 = neel_price_f.replace("-AI2980", "").replace("-AI4036", "").replace("AI", "").replace("=", "").replace("+0.01", "").split("+")
                    neel_price_f1.sort(reverse = False)
                    # print(neel_price_f1)

                    for x in neel_price_f1:
                        # print(x)
                        neel_price_fx = neel_f["AI" + str(x)].value
                        # print("neel_price_fx", neel_price_fx)
                        
                        if ":" in str(neel_price_fx):
                            neel_price_fx_s = neel_price_fx.replace("SUM", "").replace("=", "").replace("AI", "").replace(")", "").replace("(", "").replace("M", "").replace("AC", "")
                            # print(neel_price_fx_s)

                            x, y = neel_price_fx_s.split(':')
                            # print(x,y)

                            
                            for b in range(int(x) - 1, int(x)):
                                neel_child = neel["C" + str(b)].value
                                if neel_child == None:
                                    for f in range(int(x) - 3, int(x) - 2):
                                        neel_child = neel["C" + str(f)].value
                                if neel_child == "Part Number":
                                    for d in range(int(x) - 2, int(x) - 1):
                                        neel_child = neel["C" + str(d)].value
                                print(neel_child, int(x) - 2)

                            for c in range(int(x), int(y) + 1):
                                neel_sub_child = neel["C" + str(c)].value
                                if neel_sub_child == None:
                                    neel_sub_child = neel["B" + str(c)].value
                                # print(neel_sub_child)

                                neel_gross_wt = neel["E" + str(c)].value
                                if neel_gross_wt == None:
                                    neel_gross_wt = 0
                                neel_net_wt = neel["F" + str(c)].value
                                if neel_net_wt == None:
                                    neel_net_wt = 0
                                
                                if neel["T" + str(c)].value == None:
                                    neel["T" + str(c)].value = 0
                                if neel["AK" + str(c)].value == None:
                                    neel["AK" + str(c)].value = 0
                                neel_process_cost = neel["T" + str(c)].value + neel["AK" + str(c)].value
                                if neel["U" + str(c)].value == None:
                                    neel["U" + str(c)].value = 0
                                if neel["AL" + str(c)].value == None:
                                    neel["AL" + str(c)].value = 0
                                neel_dep_cost = neel["U" + str(c)].value + neel["AL" + str(c)].value
                                neel_tooling_cost = neel["V" + str(c)].value
                                if neel_tooling_cost == None:
                                    neel_tooling_cost = 0
                                neel_overhead = neel["W" + str( c)].value
                                if neel_overhead == None:
                                    neel_overhead = 0
                                neel_profit_NRMC = neel["X" + str(c)].value
                                if neel_profit_NRMC == None:
                                    neel_profit_NRMC = 0
                                neel_profit_process = neel["Y" + str(c)].value
                                if neel_profit_process == None:
                                    neel_profit_process = 0
                                neel_rej_NRMC = neel["Z" + str(c)].value
                                if neel_rej_NRMC == None:
                                    neel_rej_NRMC = 0
                                neel_rej_process = neel["AA" + str(c)].value
                                if neel_rej_process == None:
                                    neel_rej_process = 0     
                                neel_fr_NRMC = neel["AB" + str(c)].value
                                if neel_rej_process == None:
                                    neel_rej_process = 0  
                                neel_fr_process = neel["AC" + str(c)].value
                                if neel_fr_process == None:
                                    neel_fr_process = 0  
                                neel_RM_grade = neel["H" + str(c)].value
                                if neel_RM_grade == None:
                                    neel_RM_grade = " "
                                neel_no_off = neel["D" + str(c)].value
                                neel_bop_cost = neel["M" + str(c)].value

                            
                                # for hierachy

                                if neel_gross_wt == None and neel_net_wt == None and neel_process_cost == None and neel_dep_cost == None and neel_tooling_cost == None and neel_overhead == None and neel_profit_NRMC == None and neel_profit_process == None and neel_rej_NRMC == None and neel_rej_process == None and neel_fr_NRMC == None and neel_fr_process == None:
                                    continue
                                else:
                                    l1 = ["", direct_vendor, "", plan, from_date, to_date, vtv_part, "OE", neel_child, "INM"]
                                    ws.append(l1)
                                    l1 = ["", direct_vendor, "", plan, from_date, to_date, vtv_part, "OE", neel_child, "INM", neel_sub_child, "INM"]
                                    ws.append(l1)
                                    l1 = ["", direct_vendor, "", plan, from_date, to_date, vtv_part, "OE", neel_child, "INM", neel_sub_child, "INM", neel_RM_grade, "RM"]
                                    ws.append(l1)
                                    l1 = ["", direct_vendor, "", plan, from_date, to_date, vtv_part, "OE", neel_child, "INM", neel_sub_child, "INM", "Scrap", "RM"]
                                    ws.append(l1)
                                    if neel_gross_wt == None and neel_net_wt == None:
                                        l1 = ["", direct_vendor, "", plan, from_date, to_date, vtv_part, "OE", neel_child, "INM"]
                                        ws.append(l1)
                                        l1 = ["" , direct_vendor, "", plan, from_date, to_date, vtv_part, "OE", neel_child, "INM", neel_sub_child, "BOP"]
                                        ws.append(l1)
                                        print(l1)




                        elif "+" in str(neel_price_fx):
                            neel_price_fx = neel_price_fx.replace("R", "").replace("=", "").replace("AK", "").replace("AL", "").replace("AE", "").split("+")
                            # print(neel_price_fx)
                            r = neel_price_fx[0]
                            # print(r)  

                            neel_gross_wt = neel["E" + str(r)].value
                            if neel_gross_wt == None:
                                neel_gross_wt = 0
                            neel_net_wt = neel["F" + str(r)].value
                            if neel_net_wt == None:
                                neel_net_wt = 0
                            
                            if neel["T" + str(r)].value == None:
                                neel["T" + str(r)].value = 0
                            if neel["AK" + str(r)].value == None:
                                neel["AK" + str(r)].value = 0
                            neel_process_cost = neel["T" + str(r)].value + neel["AK" + str(r)].value
                            if neel["U" + str(r)].value == None:
                                neel["U" + str(r)].value = 0
                            if neel["AL" + str(r)].value == None:
                                neel["AL" + str(r)].value = 0
                            neel_dep_cost = neel["U" + str(r)].value + neel["AL" + str(r)].value
                            neel_tooling_cost = neel["V" + str(r)].value
                            if neel_tooling_cost == None:
                                neel_tooling_cost = 0
                            neel_overhead = neel["W" + str( r)].value
                            if neel_overhead == None:
                                neel_overhead = 0
                            neel_profit_NRMC = neel["X" + str(r)].value
                            if neel_profit_NRMC == None:
                                neel_profit_NRMC = 0
                            neel_profit_process = neel["Y" + str(r)].value
                            if neel_profit_process == None:
                                neel_profit_process = 0
                            neel_rej_NRMC = neel["Z" + str(r)].value
                            if neel_rej_NRMC == None:
                                neel_rej_NRMC = 0
                            neel_rej_process = neel["AA" + str(r)].value
                            if neel_rej_process == None:
                                neel_rej_process = 0
                            neel_fr_NRMC = neel["AB" + str(r)].value
                            if neel_fr_NRMC == None:
                                neel_fr_NRMC = 0
                            neel_fr_process = neel["AC" + str(r)].value
                            if neel_fr_process == None:
                                neel_fr_process = 0
                            bop = neel["C" + str(r)].value
                            if bop == None:
                                bop = neel["B" + str(r)].value
                                if bop == "BOP welding":
                                    bop = "WELD-01"
                                if bop == "Gauging Cost":
                                    bop = "GAUG-01"
                            # print(bop, r)

                            l1 = ["" , direct_vendor, "", plan, from_date, to_date, vtv_part, "OE", bop, "INM"]
                            ws.append(l1)

            ws.append({})
                               
                                

wb.save(r"F:\sequelstring\Excel\NEELMETAL_OUTPUT\Output\vtv_hierachy.xlsx")



