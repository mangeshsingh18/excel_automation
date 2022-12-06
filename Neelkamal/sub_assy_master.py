import re
import openpyxl
import psycopg2
from openpyxl import Workbook, load_workbook

#  Main Sheet
wb = Workbook()
# ws = wb.create_chartsheet
ws = openpyxl.load_workbook(r"F:\sequelstring\Excel\NEELMETAL_OUTPUT\Output\sub_assy_master.xlsx", data_only=True)
ws = wb['Sheet']

# FRT SHEET 01102021
wb1 = openpyxl.load_workbook(r"F:\sequelstring\Excel\NEELMETAL_OUTPUT\PIR_FRAME_NCR_NEELMETAL_01.01.2022.xlsx", data_only=True)
front_sheet = wb1['FRT SHEET 01102021']

wb2 = openpyxl.load_workbook(r"F:\sequelstring\Excel\NEELMETAL_OUTPUT\PIR_FRAME_NCR_NEELMETAL_01.01.2022.xlsx", data_only=False)
front_sheet_f = wb2['FRT SHEET 01102021']

# SUMMARY-FRAMEBODY
wb3 = openpyxl.load_workbook(r"F:\sequelstring\Excel\NEELMETAL_OUTPUT\PIR_FRAME_NCR_NEELMETAL_01.01.2022.xlsx", data_only=True)
summary = wb3['SUMMARY-FRAMEBODY']

wb4 = openpyxl.load_workbook(r"F:\sequelstring\Excel\NEELMETAL_OUTPUT\PIR_FRAME_NCR_NEELMETAL_01.01.2022.xlsx", data_only=False)
summary_f = wb4['SUMMARY-FRAMEBODY']

# ANNEX-VTV
wb5 = openpyxl.load_workbook(r"F:\sequelstring\Excel\NEELMETAL_OUTPUT\PIR_FRAME_NCR_NEELMETAL_01.01.2022.xlsx", data_only=True)
annex = wb5['ANNEX-VTV']

# OMAX 125cc
wb6 = openpyxl.load_workbook(r"F:\sequelstring\Excel\NEELMETAL_OUTPUT\PIR_FRAME_NCR_NEELMETAL_01.01.2022.xlsx", data_only=True)
omax = wb6['OMAX 125cc']

wb7 = openpyxl.load_workbook(r"F:\sequelstring\Excel\NEELMETAL_OUTPUT\PIR_FRAME_NCR_NEELMETAL_01.01.2022.xlsx", data_only=False)
omax_f = wb7['OMAX 125cc']

# Summary-Sub assy-Child Part
wb8 = openpyxl.load_workbook(r"F:\sequelstring\Excel\NEELMETAL_OUTPUT\PIR_FRAME_NCR_NEELMETAL_01.01.2022.xlsx", data_only=True)
sub = wb8['Summary-Sub assy-Child Part']

# AAWD-HNBL Child part Working 
wb9 = openpyxl.load_workbook(r"F:\sequelstring\Excel\NEELMETAL_OUTPUT\PIR_FRAME_NCR_NEELMETAL_01.01.2022.xlsx", data_only=True)
aawd = wb9['AAWD-HNBL Child part Working ']

wb10 = openpyxl.load_workbook(r"F:\sequelstring\Excel\NEELMETAL_OUTPUT\PIR_FRAME_NCR_NEELMETAL_01.01.2022.xlsx", data_only=False)
aawd_f = wb10['AAWD-HNBL Child part Working ']


headers = ['bom_hierarchy',	'master', 'direct_sup',	'plan',	'frequency', 'from_date',	'to_date',	'purchase_group',	'value',	'percentage',	'input_currency',	'output_currency',
        'unit',	'from_city',	'to_city',	'from_period',	'to_period',	'forward_exchange',	'leap_master',	'indicator',	'rm_exclude_flag',	'osp_conversion',	'osp_freight']

ws.append(headers)


H_column = annex["H" + str(4)].value
I_column = annex["I" + str(4)].value
J_column = annex["J" + str(4)].value
K_column = annex["K" + str(4)].value
L_column = annex["L" + str(4)].value


# for front sheet
for i in range(5,93):
    category = front_sheet["S" + str(i)].value
    # print(category)

    if category == "Sub assy & BOP":
        # print(1)
        front_partcode = front_sheet["G" + str(i)].value
        revised_price = front_sheet["P" + str(i)].value
        revised_price_f = front_sheet_f["P" + str(i)].value
        # print("front_partcode", front_partcode)
        # print("revised_price_f", revised_price_f)


        direct_vendor = front_sheet["F" + str(i)].value
        plan = front_sheet["E" + str(i)].value
        from_date = "01.01.22"
        to_date = "31.12.9999"
        # print(front_partcode, i)
        # print(revised_price_f)
        ws.append({})

        # for summary framebody
        for j in range(9, 73):
            summary_part = summary["G" + str(j)].value
            # print("summary_part", summary_part)
            summary_price = summary["BN" + str(j)].value
            summary_price_f = summary_f["BN" + str(j)].value
            summary_price1 = summary["AT" + str(j)].value

            if front_partcode == summary_part and revised_price == summary_price or revised_price == summary_price1:
                # print(summary_part)
                # print(summary_price_f)

                summary_BOP_set_f = summary_f["N" + str(j)].value
                # print(summary_BOP_set_f)

                if summary_BOP_set_f == None:
                    continue
                if "=" in str(summary_BOP_set_f):
                    # print(summary_BOP_set_f)
                    summary_BOP_set_f = re.findall("N\d+", summary_BOP_set_f)
                    # print(summary_BOP_set_f)
                    for x in summary_BOP_set_f:
                        bop_set_f = summary_f[x].value
                        # print(bop_set_f)

                        if 'ANNEX-VTV' in bop_set_f:
                            summ_vtv = summary["G" + str(j)].value
                            # print(summ_vtv,1111)

                        else:
                            # print(bop_set_f)
                            val = re.sub('[^0-9]', "", bop_set_f)
                            # print(val)
                            summ_vtv = summary["G" + str(val)].value
                            # print(summ_vtv,2222)


                        if summ_vtv == H_column: 
                            v1 = "H"
                        elif summ_vtv == I_column: 
                            v1 = "I"
                        elif summ_vtv == J_column: 
                            v1 = "J"
                        elif summ_vtv == K_column: 
                            v1 = "K"
                        elif summ_vtv == L_column: 
                            v1 = "L"

                        for a in range(6, 17):
                            no = annex[v1 + str(a)].value 
                            # print(1)
                            if no != None:
                                # print(1)
                                ann_partno = annex["C" + str(a)].value
                                # print(ann_partno)
                                po_no = annex[v1 + str(a)].value

                                l1 = [front_partcode + "_" + ann_partno, "no_off", direct_vendor, plan,"", from_date, to_date, "", po_no]
                                ws.append(l1)
                                # print(l1)
                
                summary_additional_price = summary["BL" + str(j)].value
                summary_rm_sec_operaton_charges = summary["AL" + str(20)].value
                summary_interest_cost = summary["Z" + str(j)].value
                summary_depriciation_cost = summary["AA" + str(j)].value
                summary_Wire_and_Co2_gas = summary["AB" + str(j)].value
                summary_other_cost = summary["AC" + str(j)].value
                summary_welding_fix_cost = summary["AD" + str(j)].value
                if summary_welding_fix_cost == None:
                    summary_welding_fix_cost = 0
                summary_press_shop_cost = summary["AE" + str(j)].value
                summary_OH_at_30_cost = summary["AF" + str(j)].value
                summary_profit_cost = summary["AG" + str(j)].value
                summary_rejection_at_rm_and_conv_cost = summary["AH" + str(j)].value
                summary_paking_exp_cost = summary["AI" + str(j)].value
                summary_freight_cost = summary["AJ" + str(j)].value
                summary_ammortization_cost = summary["AK" + str(j)].value
                summary_power_welding = summary["V" + str(j)].value
                summary_power_press_shop = summary["W" + str(j)].value
                summary_labour_welding = summary["R" + str(j)].value
                summary_labour_press_shop = summary["S" + str(j)].value
                summary_labour_other_assy_operation = summary["T" + str(j)].value

                
                list1 = [front_partcode, "interest_cost", direct_vendor, plan,"", from_date, to_date, "", summary_interest_cost]
                ws.append(list1)
                list1 = [front_partcode, "depreciation_cost", direct_vendor, plan,"", from_date, to_date, "", summary_depriciation_cost]
                ws.append(list1)
                list1 = [front_partcode, "Wire_and_Co2_gas", direct_vendor, plan,"", from_date, to_date, "", summary_Wire_and_Co2_gas]
                ws.append(list1)
                list1 = [front_partcode, "other_cost", direct_vendor, plan,"", from_date, to_date, "", summary_other_cost]
                ws.append(list1)
                list1 = [front_partcode, "welding_fixtures", direct_vendor, plan,"", from_date, to_date, "", summary_welding_fix_cost]
                ws.append(list1)
                list1 = [front_partcode, "power_press_shop", direct_vendor, plan,"", from_date, to_date, "", summary_press_shop_cost]
                ws.append(list1)
                list1 = [front_partcode, "OH_at_30", direct_vendor, plan,"", from_date, to_date, "", summary_OH_at_30_cost]
                ws.append(list1)
                list1 = [front_partcode, "profit", direct_vendor, plan,"", from_date, to_date, "", summary_profit_cost]
                ws.append(list1)
                list1 = [front_partcode, "rejection_on_RM_and_conv", direct_vendor, plan,"", from_date, to_date, "", summary_rejection_at_rm_and_conv_cost]
                ws.append(list1)
                list1 = [front_partcode, "packin_exp", direct_vendor, plan,"", from_date, to_date, "", summary_paking_exp_cost]
                ws.append(list1)
                list1 = [front_partcode, "fright_existing", direct_vendor, plan,"", from_date, to_date, "", summary_freight_cost]
                ws.append(list1)
                list1 = [front_partcode, "ammortization_cost", direct_vendor, plan,"", from_date, to_date, "", summary_ammortization_cost]
                ws.append(list1)
                list1 = [front_partcode, "power_welding", direct_vendor, plan,"", from_date, to_date, "", summary_power_welding]
                ws.append(list1)
                list1 = [front_partcode, "power_press_shop", direct_vendor, plan,"", from_date, to_date, "", summary_power_press_shop]
                ws.append(list1)
                list1 = [front_partcode, "labour_welding", direct_vendor, plan,"", from_date, to_date, "", summary_labour_welding]
                ws.append(list1)
                list1 = [front_partcode, "labour_press_shop", direct_vendor, plan,"", from_date, to_date, "", summary_labour_press_shop]
                ws.append(list1)
                list1 = [front_partcode, "labour_other_assy_operation", direct_vendor, plan,"", from_date, to_date, "", summary_labour_other_assy_operation]
                ws.append(list1)
                list1 = [front_partcode, "rm_sec_operaton_charges", direct_vendor, plan,"", from_date, to_date, "", summary_rm_sec_operaton_charges]
                ws.append(list1)
                list1 = [front_partcode, "additional_price", direct_vendor, plan,"", from_date, to_date, "", summary_additional_price]
                ws.append(list1)
               

                summary_bop_working_rev = summary["K" + str(j)].value
                summary_bop_working_rev_f = summary_f["K" + str(j)].value
                # print(summary_part)
                # print("summary_bop_working_rev_f", summary_bop_working_rev_f)

                if re.findall("K\d+", summary_bop_working_rev_f):
                    # print(summary_bop_working_rev_f)
                    summary_bop_working_rev_f = re.findall("K\d+", summary_bop_working_rev_f)
                    # print(summary_bop_working_rev_f)
                    for y in summary_bop_working_rev_f:
                        summary_bop_working_price_f = summary_f[y].value
                        # print(summary_bop_working_price_f)
                        part_val = re.sub("\'[A-Z\s0-9a-z]+\'\!.*|[^0-9]", "",str(summary_bop_working_price_f))
                        # print(part_val)
                        summary_bop_working_part = summary["G" + str(part_val)].value
                        # print(summary_bop_working_part)
                        summary_bop_working_price = summary[y].value
                        # print(summary_bop_working_price)
                

                else:
                    # print(summary_bop_working_rev_f)
                    summary_bop_working_rev_f = re.findall("K\$\d+", summary_bop_working_rev_f)
                    # summary_bop_working_rev_f = re.sub("\$","",  str(summary_bop_working_rev_f)).strip()
                    # print(summary_bop_working_rev_f)
                    for y in summary_bop_working_rev_f:
                        summary_bop_working_price_f = summary_f[y].value
                        # print(summary_bop_working_price_f)
                        val = re.sub("I\$\d+|[^0-9]","", str(summary_bop_working_price_f))
                        # print(val)
                        
                        summary_bop_working_price_f = summary_f["K" +str(val)].value
                        # print(summary_bop_working_price_f)
                        summary_bop_working_part = summary["G" + str(val)].value
                        # print(summary_bop_working_part)
                        summary_bop_working_price = summary["K" +str(val)].value
                        # print(summary_bop_working_price)

                # for 'OMAX 125cc'
                for o in range(33, 760):
                    omax_part = omax["C" + str(o)].value
                    omax_price = omax["AI" + str(o)].value
                    omax_price_f = omax_f["AI" + str(o)].value
                    # print(omax_price_f)
                    if omax_part == summary_bop_working_part and summary_bop_working_price == omax_price:
                        # print(omax_part)
                        # print(omax_price_f)
                        omax_price_f = omax_price_f.replace("AI", "").replace("=", "").replace("+", "")
                        omax_price_f = omax_f["AI" + str(omax_price_f)].value
                        omax_price_f1 = omax_price_f.replace("AI", "").replace("=", "").split("+")
                        omax_price_f1 = omax_price_f1[::-1]
                        # print(omax_price_f1)
                        for x in omax_price_f1:
                            # print(x)
                            omax_price_fx = omax_f["AI" + str(x)].value
                            # print(omax_price_fx ,"AI" + str(x))

                            if ":" in omax_price_fx:
                                omax_price_fx_s = omax_price_fx.replace("SUM", "").replace("=", "").replace("AI", "").replace(")", "").replace("(", "")
                                # print(omax_price_fx_s)

                                x, y = omax_price_fx_s.split(':')
                                # print(x,y)

                                for b in range(int(x) - 1, int(x)):
                                    omax_child = omax["C" + str(b)].value
                                    if omax_child == None:
                                        continue
                                    elif omax_child == "Part Number":
                                        for d in range(int(x) - 2, int(x) - 1):
                                            omax_child = omax["C" + str(d)].value
                                    # print(omax_child, b)

                                for c in range(int(x), int(y) + 1):
                                    omax_sub_child = omax["C" + str(c)].value
                                    if omax_sub_child == None:
                                        omax_sub_child = omax["B" + str(c)].value
                                    # print(omax_sub_child)

                                    omax_nrmc = omax["R" + str(c)].value
                                    omax_gross_wt = omax["E" + str(c)].value
                                    if omax_gross_wt == None:
                                        omax_gross_wt = 0 
                                    omax_net_wt = omax["F" + str(c)].value
                                    if omax_net_wt == None:
                                        omax_net_wt = 0 
                                    omax_scrap_wt = omax["G" + str(c)].value
                                    if omax_scrap_wt == None:
                                        omax_scrap_wt = 0 
                                    omax_rm_grade = omax["H" + str(c)].value
                                    if omax_rm_grade == None:
                                        omax_rm_grade = 0 
                                    omax_process_cost = omax["T" + str(c)].value + omax["AK" + str(c)].value
                                    # print(omax["T" + str(c)].value)
                                    # print(omax["AK" + str(c)].value)
                                    if omax["U" + str(c)].value == None:
                                        omax["U" + str(c)].value = 0
                                    if omax["AL" + str(c)].value == None:
                                        omax["AL" + str(c)].value = 0
                                    omax_dep_cost = omax["U" + str(c)].value + omax["AL" + str(c)].value
                                    omax_tooling_cost = omax["V" + str(c)].value
                                    if omax_tooling_cost == None:
                                        omax_tooling_cost = 0 
                                    omax_overhead = omax["W" + str( c)].value
                                    if omax_overhead == None:
                                        omax_overhead = 0 
                                    omax_profit_NRMC = omax["X" + str(c)].value
                                    if omax_profit_NRMC == None:
                                        omax_profit_NRMC = 0 
                                    omax_profit_process = omax["Y" + str(c)].value
                                    if omax_profit_process == None:
                                        omax_profit_process = 0 
                                    omax_rej_NRMC = omax["Z" + str(c)].value
                                    if omax_rej_NRMC == None:
                                        omax_rej_NRMC = 0 
                                    omax_rej_process = omax["AA" + str(c)].value
                                    if omax_rej_process == None:
                                        omax_rej_process = 0 
                                    omax_fr_NRMC = omax["AB" + str(c)].value
                                    if omax_fr_NRMC == None:
                                        omax_fr_NRMC = 0 
                                    omax_fr_process = omax["AC" + str(c)].value
                                    if omax_fr_process == None:
                                        omax_fr_process = 0 

                                    omax_no_off = omax["D" + str(c)].value
                                    omax_bop_cost = omax["M" + str(c)].value

                                    if omax_gross_wt == None and omax_net_wt == None and omax_process_cost == None and omax_dep_cost == None and omax_tooling_cost == None and omax_overhead == None and omax_profit_NRMC == None and omax_profit_process == None and omax_rej_NRMC == None and omax_rej_process == None and omax_fr_NRMC == None and omax_fr_process == None:
                                        continue
                                    else:
                                        l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child) + "_" + str(omax_rm_grade), "gross_wt", direct_vendor, plan,"", from_date, to_date, "", omax_gross_wt]
                                        ws.append(l1)
                                        l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child) + "_" + str(omax_rm_grade), "net_wt", direct_vendor, plan,"", from_date, to_date, "", omax_net_wt]
                                        ws.append(l1)
                                        l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child), "process_cost", direct_vendor, plan,"", from_date, to_date, "", omax_process_cost]
                                        ws.append(l1)
                                        l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child), "dep_cost", direct_vendor, plan,"", from_date, to_date, "", omax_dep_cost]
                                        ws.append(l1)
                                        l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child), "tooling_cost", direct_vendor, plan,"", from_date, to_date, "", omax_tooling_cost]
                                        ws.append(l1)
                                        l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child), "overhead", direct_vendor, plan,"", from_date, to_date, "", omax_overhead]
                                        ws.append(l1)
                                        l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child), "profit_NRMC", direct_vendor, plan,"", from_date, to_date, "", omax_profit_NRMC]
                                        ws.append(l1)
                                        l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child), "profit_process", direct_vendor, plan,"", from_date, to_date, "", omax_profit_process]
                                        ws.append(l1)
                                        l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child), "rej_NRMC", direct_vendor, plan,"", from_date, to_date, "", omax_rej_NRMC]
                                        ws.append(l1)
                                        l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child), "rej_process", direct_vendor, plan,"", from_date, to_date, "", omax_rej_process]
                                        ws.append(l1)
                                        l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child), "fr_NRMC", direct_vendor, plan,"", from_date, to_date, "", omax_fr_NRMC]
                                        ws.append(l1)
                                        l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child), "fr_process", direct_vendor, plan,"", from_date, to_date, "", omax_fr_process]
                                        ws.append(l1)
                                
                                        if omax_gross_wt == None and omax_net_wt == None:
                                            l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child), "NO_OFF", direct_vendor, plan,"", from_date, to_date, "", omax_no_off]
                                            ws.append(l1)
                                            l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child), "BOP_COST", direct_vendor, plan,"", from_date, to_date, "", omax_bop_cost]
                                            ws.append(l1)

                            elif "+" in omax_price_fx:
                                if omax_price_fx[:2] == "=R":
                                    # print(omax_price_fx, "ncjdnjcvd")
                                    omax_price_fx_r = omax_price_fx.replace("R", "").replace("=", "").replace("AK", "").replace("AL", "").replace("AE", "").split("+")
                                    r = omax_price_fx_r[0]
                                    # print(r)
                                    # print(omax_price_fx_r)

                                    omax_gross_wt = omax["E" + str(r)].value
                                    if omax_gross_wt == None:
                                            omax_gross_wt = 0 
                                        # print(omax_gross_wt,"E" + str(r))
                                    omax_net_wt = omax["F" + str(r)].value
                                    if omax_net_wt == None:
                                        omax_net_wt = 0 
                                    omax_scrap_wt = omax["G" + str(r)].value
                                    if omax_scrap_wt == None:
                                        omax_scrap_wt = 0 
                                    omax_rm_grade = omax["H" + str(r)].value
                                    if omax_rm_grade == None:
                                        omax_rm_grade = 0 
                                    omax_revised_NRMC = omax["R" + str(r)].value
                                    if omax_revised_NRMC == None:
                                        omax_revised_NRMC = 0 
                                    if omax["AK" + str(r)].value == None : 
                                        omax["AK" + str(r)].value = 0
                                    if omax["T" + str(r)].value == None:
                                        omax["T" + str(r)].value = 0
                                    omax_process_cost = omax["T" + str(r)].value + omax["AK" + str(r)].value
                                        # print(omax_process_cost)
                            
                                    # print(omax["T" + str(r)].value)
                                    # print(omax["AK" + str(r)].value)
                                    if omax["U" + str(r)].value == None: 
                                        omax["U" + str(r)].value = 0
                                    if omax["AL" + str(r)].value == None:
                                        omax["AL" + str(r)].value = 0
                                    omax_dep_cost = omax["U" + str(r)].value + omax["AL" + str(r)].value
                                    # print(omax_dep_cost)
                                    omax_tooling_cost = omax["V" + str(r)].value
                                    if omax_tooling_cost == None:
                                        omax_tooling_cost = 0 
                                    omax_overhead = omax["W" + str(r)].value
                                    if omax_overhead == None:
                                        omax_overhead = 0 
                                    omax_profit_NRMC = omax["X" + str(r)].value
                                    if omax_profit_NRMC == None:
                                        omax_profit_NRMC = 0 
                                    omax_profit_process = omax["Y" + str(r)].value
                                    if omax_profit_process == None:
                                        omax_profit_process = 0 
                                    omax_rej_NRMC = omax["Z" + str(r)].value
                                    if omax_rej_NRMC == None:
                                        omax_rej_NRMC = 0 
                                    omax_rej_process = omax["AA" + str(r)].value
                                    if omax_rej_process == None:
                                        omax_rej_process = 0 
                                    omax_fr_NRMC = omax["AB" + str(r)].value
                                    if omax_fr_NRMC == None:
                                        omax_fr_NRMC = 0 
                                    omax_fr_process = omax["AC" + str(r)].value
                                    if omax_fr_process == None:
                                        omax_fr_process = 0 
                                    omax_revised_total_cost = omax["AI" + str(r)].value
                                    if omax_revised_total_cost == None:
                                        omax_revised_total_cost = 0 
                                    omax_rm_grade = omax["H" + str(r)].value
                                    if omax_rm_grade == None:
                                        omax_rm_grade = " "
                                    # print(omax_rm_grade)
                                    # print(omax_revised_total_cost, x)
                                    omax_NRMC = omax["R" + str(r)].value
                                    if omax_NRMC == None:
                                        omax_NRMC = 0 
                                    # print(omax_NRMC, r)
                                    bop = omax["C" + str(r)].value
                                    if bop == None:
                                        bop = omax["B" + str(r)].value
                                        if bop == "BOP welding":
                                            bop = "WELD-01"
                                        if bop == "Gauging Cost":
                                            bop = "GAUG-01"
                                    # print(bop, r)

                                    if omax_gross_wt == None and omax_net_wt == None and omax_process_cost == None and omax_dep_cost == None and omax_tooling_cost == None and omax_overhead == None and omax_profit_NRMC == None and omax_profit_process == None and omax_rej_NRMC == None and omax_rej_process == None and omax_fr_NRMC == None and omax_fr_process == None:
                                        continue
                                    else:
                                        l1 = [front_partcode + "_" + bop, "gross_wt", direct_vendor, plan,"", from_date, to_date, "", omax_gross_wt]
                                        ws.append(l1)
                                        print(l1, "222222222222222222")
                                        l1 = [front_partcode + "_" + bop , "net_wt", direct_vendor, plan,"", from_date, to_date, "", omax_net_wt]
                                        ws.append(l1)
                                        l1 = [front_partcode + "_" + bop, "process_cost", direct_vendor, plan,"", from_date, to_date, "", omax_process_cost]
                                        ws.append(l1)
                                        l1 = [front_partcode + "_" + bop, "dep_cost", direct_vendor, plan,"", from_date, to_date, "", omax_dep_cost]
                                        ws.append(l1)
                                        l1 = [front_partcode + "_" + bop, "tooling_cost", direct_vendor, plan,"", from_date, to_date, "", omax_tooling_cost]
                                        ws.append(l1)
                                        l1 = [front_partcode + "_" + bop, "overhead", direct_vendor, plan,"", from_date, to_date, "", omax_overhead]
                                        ws.append(l1)
                                        l1 = [front_partcode + "_" + bop, "profit_NRMC", direct_vendor, plan,"", from_date, to_date, "", omax_profit_NRMC]
                                        ws.append(l1)
                                        l1 = [front_partcode + "_" + bop, "profit_process", direct_vendor, plan,"", from_date, to_date, "", omax_profit_process]
                                        ws.append(l1)
                                        l1 = [front_partcode + "_" + bop, "rej_NRMC", direct_vendor, plan,"", from_date, to_date, "", omax_rej_NRMC]
                                        ws.append(l1)
                                        l1 = [front_partcode + "_" + bop, "rej_process", direct_vendor, plan,"", from_date, to_date, "", omax_rej_process]
                                        ws.append(l1)
                                        l1 = [front_partcode + "_" + bop, "fr_NRMC", direct_vendor, plan,"", from_date, to_date, "", omax_fr_NRMC]
                                        ws.append(l1)
                                        l1 = [front_partcode + "_" + bop, "fr_process", direct_vendor, plan,"", from_date, to_date, "", omax_fr_process]
                                        ws.append(l1)


        # Summary-Sub assy-Child Part
        for s in range(11, 36):
            sub_part = sub["D" + str(s)].value
            sub_price = sub["AU" + str(s)].value
            if front_partcode == sub_part and revised_price == sub_price:
                # print("sub_part", sub_part)
                # print("sub_price", sub_price)

                sub_interest_cost = sub["X" + str(s)].value
                if sub_interest_cost == None:
                    sub_interest_cost = 0
                sub_depriciation_cost = sub["Y" + str(s)].value
                if sub_depriciation_cost == None:
                    sub_depriciation_cost = 0
                sub_Wire_and_Co2_gas = sub["Z" + str(s)].value
                if sub_Wire_and_Co2_gas == None:
                    sub_Wire_and_Co2_gas = 0
                sub_other_cost = sub["AA" + str(s)].value
                if sub_other_cost == None:
                    sub_other_cost = 0
                sub_welding_fix_cost = sub["AB" + str(s)].value
                if sub_welding_fix_cost == None:
                    sub_welding_fix_cost = 0
                sub_press_shop_cost = sub["AC" + str(s)].value
                if sub_press_shop_cost == None:
                    sub_press_shop_cost = 0
                sub_OH_at_30_cost = sub["AD" + str(s)].value
                sub_profit_cost = sub["AE" + str(s)].value
                sub_rejection_at_rm_and_conv_cost = sub["AF" + str(s)].value
                sub_paking_exp_cost = sub["AG" + str(s)].value
                sub_freight_cost = sub["AH" + str(s)].value
                sub_ammortization_cost = sub["AI" + str(s)].value
                if sub_ammortization_cost == None:
                     sub_ammortization_cost = 0
                sub_total = sub["AJ" + str(s)].value
                if sub_total == None:
                     sub_total = 0                  
                sub_additional_price = sub["BL" + str(s)].value
                if sub_additional_price == None:
                    sub_additional_price = 0
                sub_rm_sec_operaton_charges = sub["AK" + str(s)].value
                sub_delta_due_to_labour_on_welding = sub["AL" + str(s)].value
                sub_delta_due_to_labour_on_press_shop = sub["AM" + str(s)].value
                sub_delta_due_to_labour_on_other_ops = sub["AN" + str(s)].value
                sub_delta_due_to_interest_on_press_shop = sub["AO" + str(s)].value
                if sub_delta_due_to_interest_on_press_shop == None:
                     sub_delta_due_to_interest_on_press_shop = 0
                sub_welding_cost_revised = sub["K" + str(s)].value
                if sub_welding_cost_revised == None:
                    sub_welding_cost_revised = 0
                sub_labour = sub["O" + str(s)].value
                if sub_labour == None:
                    sub_labour = 0
                sub_power = sub["S" + str(s)].value
                if sub_power == None:
                    sub_power = 0
                sub_welding_cost_revised = sub["H" + str(s)].value
                if sub_welding_cost_revised == None:
                    sub_welding_cost_revised = 0

                # AAWD-HNBL Child part Working 
                for a in range(29, 209):
                    aawd_part = aawd["D" + str(a)].value
                    aawd_price = aawd["AL" + str(a)].value

                    if aawd_part == sub_part and aawd_price == sub_welding_cost_revised:
                        aawd_price_f = aawd_f["AL" + str(a)].value
                        # print(aawd_price_f)
                        if ":" in aawd_price_f:
                            aawd_price_f_s = aawd_price_f.replace("SUM", "").replace("=", "").replace("AL", "").replace(")", "").replace("(", "")
                            # print(aawd_price_f_s)
                            x, y = aawd_price_f_s.split(':')
                            # print(x,y)

                            for b in range(int(x) - 1, int(x)):
                                aawd_child = aawd["B" + str(b)].value
                                # print(aawd_child, b)
                            
                            for c in range(int(x), int(y) + 1):
                                aawd_sub_child = aawd["C" + str(c)].value
                                # print(aawd_sub_child)

                                # aawd_process_cost_conversion_inc = aawd["AN" + str(a)].value
                                # # print(aawd_process_cost_conversion_inc)
                                # aawd_int_rationalisation = aawd["AO" + str(a)].value

                                
                                aawd_gross_wt = aawd["H" + str(c)].value
                                if aawd_gross_wt == None:
                                    aawd_gross_wt = 0 
                                aawd_net_wt = aawd["I" + str(c)].value
                                if aawd_net_wt == None:
                                    aawd_net_wt = 0 
                                aawd_scrap_wt = aawd["J" + str(c)].value
                                if aawd_scrap_wt == None:
                                    aawd_scrap_wt = 0 
                                aawd_rm_grade = aawd["K" + str(c)].value
                                if aawd_rm_grade == None:
                                    aawd_rm_grade = 0 

                                if aawd["W" + str(c)].value == None:
                                    aawd["W" + str(c)].value = 0
                                if aawd["AN" + str(c)].value == None:
                                    aawd["AN" + str(c)].value = 0
                                aawd_process_cost = aawd["W" + str(c)].value + aawd["AN" + str(c)].value
                                if aawd["X" + str(c)].value == None:
                                    aawd["X" + str(c)].value = 0
                                if aawd["AO" + str(c)].value == None:
                                    aawd["AO" + str(c)].value = 0
                                aawd_dep_cost = aawd["X" + str(c)].value + aawd["AO" + str(c)].value
                                aawd_tooling_cost = aawd["Y" + str(c)].value
                                if aawd_tooling_cost == None:
                                    aawd_tooling_cost = 0 
                                aawd_overhead = aawd["Z" + str( c)].value
                                if aawd_overhead == None:
                                    aawd_overhead = 0 
                                aawd_profit_NRMC = aawd["AA" + str(c)].value
                                if aawd_profit_NRMC == None:
                                    aawd_profit_NRMC = 0 
                                aawd_profit_process = aawd["AB" + str(c)].value
                                if aawd_profit_process == None:
                                    aawd_profit_process = 0 
                                aawd_rej_NRMC = aawd["AC" + str(c)].value
                                if aawd_rej_NRMC == None:
                                    aawd_rej_NRMC = 0 
                                aawd_rej_process = aawd["AD" + str(c)].value
                                if aawd_rej_process == None:
                                    aawd_rej_process = 0 
                                aawd_fr_NRMC = aawd["AE" + str(c)].value
                                if aawd_fr_NRMC == None:
                                    aawd_fr_NRMC = 0 
                                aawd_fr_process = aawd["AF" + str(c)].value
                                if aawd_fr_process == None:
                                    aawd_fr_process = 0 

                                aawd_no_off = aawd["G" + str(c)].value
                                aawd_bop_cost = aawd["R" + str(c)].value

                                if aawd_gross_wt == None and aawd_net_wt == None and aawd_process_cost == None and aawd_dep_cost == None and aawd_tooling_cost == None and aawd_overhead == None and aawd_profit_NRMC == None and aawd_profit_process == None and aawd_rej_NRMC == None and aawd_rej_process == None and aawd_fr_NRMC == None and aawd_fr_process == None:
                                    continue
                                else: 
                                    l1 = [str(front_partcode) + "_" + str(aawd_child) + "_" + str(aawd_sub_child) + "_" + str(aawd_rm_grade), "gross_wt", direct_vendor, plan,"", from_date, to_date, "", aawd_gross_wt]
                                    ws.append(l1)
                                    l1 = [str(front_partcode) + "_" + str(aawd_child) + "_" + str(aawd_sub_child) + "_" + str(aawd_rm_grade), "net_wt", direct_vendor, plan,"", from_date, to_date, "", aawd_net_wt]
                                    ws.append(l1)
                                    l1 = [str(front_partcode) + "_" + str(aawd_child) + "_" + str(aawd_sub_child), "process_cost", direct_vendor, plan,"", from_date, to_date, "", aawd_process_cost]
                                    ws.append(l1)
                                    l1 = [str(front_partcode) + "_" + str(aawd_child) + "_" + str(aawd_sub_child), "dep_cost", direct_vendor, plan,"", from_date, to_date, "", aawd_dep_cost]
                                    ws.append(l1)
                                    l1 = [str(front_partcode) + "_" + str(aawd_child) + "_" + str(aawd_sub_child), "tooling_cost", direct_vendor, plan,"", from_date, to_date, "", aawd_tooling_cost]
                                    ws.append(l1)
                                    l1 = [str(front_partcode) + "_" + str(aawd_child) + "_" + str(aawd_sub_child), "overhead", direct_vendor, plan,"", from_date, to_date, "", aawd_overhead]
                                    ws.append(l1)
                                    l1 = [str(front_partcode) + "_" + str(aawd_child) + "_" + str(aawd_sub_child), "profit_NRMC", direct_vendor, plan,"", from_date, to_date, "", aawd_profit_NRMC]
                                    ws.append(l1)
                                    l1 = [str(front_partcode) + "_" + str(aawd_child) + "_" + str(aawd_sub_child), "profit_process", direct_vendor, plan,"", from_date, to_date, "", aawd_profit_process]
                                    ws.append(l1)
                                    l1 = [str(front_partcode) + "_" + str(aawd_child) + "_" + str(aawd_sub_child), "rej_NRMC", direct_vendor, plan,"", from_date, to_date, "", aawd_rej_NRMC]
                                    ws.append(l1)
                                    l1 = [str(front_partcode) + "_" + str(aawd_child) + "_" + str(aawd_sub_child), "rej_process", direct_vendor, plan,"", from_date, to_date, "", aawd_rej_process]
                                    ws.append(l1)
                                    l1 = [str(front_partcode) + "_" + str(aawd_child) + "_" + str(aawd_sub_child), "fr_NRMC", direct_vendor, plan,"", from_date, to_date, "", aawd_fr_NRMC]
                                    ws.append(l1)
                                    l1 = [str(front_partcode) + "_" + str(aawd_child) + "_" + str(aawd_sub_child), "fr_process", direct_vendor, plan,"", from_date, to_date, "", aawd_fr_process]
                                    ws.append(l1)
                                    
                                    if aawd_gross_wt == None and aawd_net_wt == None:
                                        l1 = [str(front_partcode) + "_" + str(aawd_child) + "_" + str(aawd_sub_child), "NO_OFF", direct_vendor, plan,"", from_date, to_date, "", aawd_no_off]
                                        ws.append(l1)
                                        l1 = [str(front_partcode) + "_" + str(aawd_child) + "_" + str(aawd_sub_child), "BOP_COST", direct_vendor, plan,"", from_date, to_date, "", aawd_bop_cost]
                                        ws.append(l1)




wb.save(r"F:\sequelstring\Excel\NEELMETAL_OUTPUT\Output\sub_assy_master.xlsx")






                        

                    







#                     if 'K$18' in summary_bop_working_rev_f:
#                         # print(summary_bop_working_rev_f)
#                         val = re.sub("I\d+.*|[^0-9]", "", summary_bop_working_rev_f)
#                         # print(val)
#                         # print(summary_part)
#                         summary_bop_rev = summary_f["K" + str(val)].value
#                         # print(summary_bop_rev)
#                         if "K" in summary_bop_rev:
#                             v = re.sub("I\$\d+|[^0-9]", "", summary_bop_rev)
#                             summary_col = summary["K" + str(v)].value
#                             summary_part_col = summary["G" + str(v)].value
#                             # print(summary_col)
#                             # print(summary_part)

#                             # for 'OMAX 125cc'
#                             for o in range(33, 760):
#                                 omax_part = omax["C" + str(o)].value
#                                 omax_price = omax["AI" + str(o)].value
#                                 omax_price_f = omax_f["AI" + str(o)].value
#                                 # print(omax_price_f)

                                # if omax_part == summary_part_col and summary_col == omax_price or omax_part == summary_part_col and summary_bop_working_rev == omax_price:
                                #     # print(omax_part)
                                #     # print(omax_price_f)

#                                     omax_price_f = omax_price_f.replace("AI", "").replace("=", "").replace("+", "")
#                                     omax_price_f = omax_f["AI" + str(omax_price_f)].value
#                                     omax_price_f1 = omax_price_f.replace("AI", "").replace("=", "").split("+")
#                                     omax_price_f1.sort(reverse = False)
#                                     # print(omax_price_f1)

#                                     for x in omax_price_f1:
#                                         # print(x)
#                                         omax_price_fx = omax_f["AI" + str(x)].value
#                                         # print(omax_price_fx ,"AI" + str(x))

#                                         if ":" in omax_price_fx:
#                                             omax_price_fx_s = omax_price_fx.replace("SUM", "").replace("=", "").replace("AI", "").replace(")", "").replace("(", "")
#                                             # print(omax_price_fx_s)

#                                             x, y = omax_price_fx_s.split(':')
#                                             # print(x,y)

#                                             for b in range(int(x) - 1, int(x)):
#                                                 omax_child = omax["C" + str(b)].value
#                                                 if omax_child == None:
#                                                     continue
#                                                 elif omax_child == "Part Number":
#                                                     for d in range(int(x) - 2, int(x) - 1):
#                                                         omax_child = omax["C" + str(d)].value
#                                                 # print(omax_child, b)

#                                             for c in range(int(x), int(y) + 1):
#                                                 omax_sub_child = omax["C" + str(c)].value
#                                                 if omax_sub_child == None:
#                                                     omax_sub_child = omax["B" + str(c)].value
#                                                 # print(omax_sub_child)

#                                                 omax_nrmc = omax["R" + str(c)].value
#                                                 omax_gross_wt = omax["E" + str(c)].value
#                                                 if omax_gross_wt == None:
#                                                     omax_gross_wt = 0 
#                                                 omax_net_wt = omax["F" + str(c)].value
#                                                 if omax_net_wt == None:
#                                                     omax_net_wt = 0 
#                                                 omax_scrap_wt = omax["G" + str(c)].value
#                                                 if omax_scrap_wt == None:
#                                                     omax_scrap_wt = 0 
#                                                 omax_rm_grade = omax["H" + str(c)].value
#                                                 if omax_rm_grade == None:
#                                                     omax_rm_grade = 0 
#                                                 omax_process_cost = omax["T" + str(c)].value + omax["AK" + str(c)].value
#                                                 # print(omax["T" + str(c)].value)
#                                                 # print(omax["AK" + str(c)].value)
#                                                 if omax["U" + str(c)].value == None:
#                                                     omax["U" + str(c)].value = 0
#                                                 if omax["AL" + str(c)].value == None:
#                                                     omax["AL" + str(c)].value = 0
#                                                 omax_dep_cost = omax["U" + str(c)].value + omax["AL" + str(c)].value
#                                                 omax_tooling_cost = omax["V" + str(c)].value
#                                                 if omax_tooling_cost == None:
#                                                     omax_tooling_cost = 0 
#                                                 omax_overhead = omax["W" + str( c)].value
#                                                 if omax_overhead == None:
#                                                     omax_overhead = 0 
#                                                 omax_profit_NRMC = omax["X" + str(c)].value
#                                                 if omax_profit_NRMC == None:
#                                                     omax_profit_NRMC = 0 
#                                                 omax_profit_process = omax["Y" + str(c)].value
#                                                 if omax_profit_process == None:
#                                                     omax_profit_process = 0 
#                                                 omax_rej_NRMC = omax["Z" + str(c)].value
#                                                 if omax_rej_NRMC == None:
#                                                     omax_rej_NRMC = 0 
#                                                 omax_rej_process = omax["AA" + str(c)].value
#                                                 if omax_rej_process == None:
#                                                     omax_rej_process = 0 
#                                                 omax_fr_NRMC = omax["AB" + str(c)].value
#                                                 if omax_fr_NRMC == None:
#                                                     omax_fr_NRMC = 0 
#                                                 omax_fr_process = omax["AC" + str(c)].value
#                                                 if omax_fr_process == None:
#                                                     omax_fr_process = 0 

#                                                 omax_no_off = omax["D" + str(c)].value
#                                                 omax_bop_cost = omax["M" + str(c)].value

#                                                 if omax_gross_wt == None and omax_net_wt == None and omax_process_cost == None and omax_dep_cost == None and omax_tooling_cost == None and omax_overhead == None and omax_profit_NRMC == None and omax_profit_process == None and omax_rej_NRMC == None and omax_rej_process == None and omax_fr_NRMC == None and omax_fr_process == None:
#                                                     continue
#                                                 else:
#                                                     l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child) + "_" + str(omax_rm_grade), "gross_wt", direct_vendor, plan,"", from_date, to_date, "", omax_gross_wt]
#                                                     ws.append(l1)
#                                                     l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child) + "_" + str(omax_rm_grade), "net_wt", direct_vendor, plan,"", from_date, to_date, "", omax_net_wt]
#                                                     ws.append(l1)
#                                                     l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child), "process_cost", direct_vendor, plan,"", from_date, to_date, "", omax_process_cost]
#                                                     ws.append(l1)
#                                                     l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child), "dep_cost", direct_vendor, plan,"", from_date, to_date, "", omax_dep_cost]
#                                                     ws.append(l1)
#                                                     l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child), "tooling_cost", direct_vendor, plan,"", from_date, to_date, "", omax_tooling_cost]
#                                                     ws.append(l1)
#                                                     l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child), "overhead", direct_vendor, plan,"", from_date, to_date, "", omax_overhead]
#                                                     ws.append(l1)
#                                                     l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child), "profit_NRMC", direct_vendor, plan,"", from_date, to_date, "", omax_profit_NRMC]
#                                                     ws.append(l1)
#                                                     l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child), "profit_process", direct_vendor, plan,"", from_date, to_date, "", omax_profit_process]
#                                                     ws.append(l1)
#                                                     l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child), "rej_NRMC", direct_vendor, plan,"", from_date, to_date, "", omax_rej_NRMC]
#                                                     ws.append(l1)
#                                                     l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child), "rej_process", direct_vendor, plan,"", from_date, to_date, "", omax_rej_process]
#                                                     ws.append(l1)
#                                                     l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child), "fr_NRMC", direct_vendor, plan,"", from_date, to_date, "", omax_fr_NRMC]
#                                                     ws.append(l1)
#                                                     l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child), "fr_process", direct_vendor, plan,"", from_date, to_date, "", omax_fr_process]
#                                                     ws.append(l1)
                                                    
#                                                     if omax_gross_wt == None and omax_net_wt == None:
#                                                         l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child), "NO_OFF", direct_vendor, plan,"", from_date, to_date, "", omax_no_off]
#                                                         ws.append(l1)
#                                                         l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child), "BOP_COST", direct_vendor, plan,"", from_date, to_date, "", omax_bop_cost]
#                                                         ws.append(l1)

#                                         elif "+" in omax_price_fx:
#                                             if omax_price_fx[:2] == "=R":
#                                                 # print(omax_price_fx, "ncjdnjcvd")
#                                                 omax_price_fx_r = omax_price_fx.replace("R", "").replace("=", "").replace("AK", "").replace("AL", "").replace("AE", "").split("+")
#                                                 r = omax_price_fx_r[0]
#                                                 # print(r)
#                                                 # print(omax_price_fx_r)

#                                                 omax_gross_wt = omax["E" + str(r)].value
#                                                 if omax_gross_wt == None:
#                                                         omax_gross_wt = 0 
#                                                     # print(omax_gross_wt,"E" + str(r))
#                                                 omax_net_wt = omax["F" + str(r)].value
#                                                 if omax_net_wt == None:
#                                                     omax_net_wt = 0 
#                                                 omax_scrap_wt = omax["G" + str(r)].value
#                                                 if omax_scrap_wt == None:
#                                                     omax_scrap_wt = 0 
#                                                 omax_rm_grade = omax["H" + str(r)].value
#                                                 if omax_rm_grade == None:
#                                                     omax_rm_grade = 0 
#                                                 omax_revised_NRMC = omax["R" + str(r)].value
#                                                 if omax_revised_NRMC == None:
#                                                     omax_revised_NRMC = 0 
#                                                 if omax["AK" + str(r)].value == None : 
#                                                     omax["AK" + str(r)].value = 0
#                                                 if omax["T" + str(r)].value == None:
#                                                     omax["T" + str(r)].value = 0
#                                                 omax_process_cost = omax["T" + str(r)].value + omax["AK" + str(r)].value
#                                                     # print(omax_process_cost)
                                                
#                                                 # print(omax["T" + str(r)].value)
#                                                 # print(omax["AK" + str(r)].value)
#                                                 if omax["U" + str(r)].value == None: 
#                                                     omax["U" + str(r)].value = 0
#                                                 if omax["AL" + str(r)].value == None:
#                                                     omax["AL" + str(r)].value = 0
#                                                 omax_dep_cost = omax["U" + str(r)].value + omax["AL" + str(r)].value
#                                                 # print(omax_dep_cost)
#                                                 omax_tooling_cost = omax["V" + str(r)].value
#                                                 if omax_tooling_cost == None:
#                                                     omax_tooling_cost = 0 
#                                                 omax_overhead = omax["W" + str(r)].value
#                                                 if omax_overhead == None:
#                                                     omax_overhead = 0 
#                                                 omax_profit_NRMC = omax["X" + str(r)].value
#                                                 if omax_profit_NRMC == None:
#                                                     omax_profit_NRMC = 0 
#                                                 omax_profit_process = omax["Y" + str(r)].value
#                                                 if omax_profit_process == None:
#                                                     omax_profit_process = 0 
#                                                 omax_rej_NRMC = omax["Z" + str(r)].value
#                                                 if omax_rej_NRMC == None:
#                                                     omax_rej_NRMC = 0 
#                                                 omax_rej_process = omax["AA" + str(r)].value
#                                                 if omax_rej_process == None:
#                                                     omax_rej_process = 0 
#                                                 omax_fr_NRMC = omax["AB" + str(r)].value
#                                                 if omax_fr_NRMC == None:
#                                                     omax_fr_NRMC = 0 
#                                                 omax_fr_process = omax["AC" + str(r)].value
#                                                 if omax_fr_process == None:
#                                                     omax_fr_process = 0 
#                                                 omax_revised_total_cost = omax["AI" + str(r)].value
#                                                 if omax_revised_total_cost == None:
#                                                     omax_revised_total_cost = 0 
#                                                 omax_rm_grade = omax["H" + str(r)].value
#                                                 if omax_rm_grade == None:
#                                                     omax_rm_grade = " "
#                                                 # print(omax_rm_grade)
#                                                 # print(omax_revised_total_cost, x)
#                                                 omax_NRMC = omax["R" + str(r)].value
#                                                 if omax_NRMC == None:
#                                                     omax_NRMC = 0 
#                                                 # print(omax_NRMC, r)
#                                                 bop = omax["C" + str(r)].value
#                                                 if bop == None:
#                                                     bop = omax["B" + str(r)].value
#                                                     if bop == "BOP welding":
#                                                         bop = "WELD-01"
#                                                     if bop == "Gauging Cost":
#                                                         bop = "GAUG-01"
#                                                 # print(bop, r)

#                                                 if omax_gross_wt == None and omax_net_wt == None and omax_process_cost == None and omax_dep_cost == None and omax_tooling_cost == None and omax_overhead == None and omax_profit_NRMC == None and omax_profit_process == None and omax_rej_NRMC == None and omax_rej_process == None and omax_fr_NRMC == None and omax_fr_process == None:
#                                                     continue
#                                                 else:
#                                                     l1 = [front_partcode + "_" + bop, "gross_wt", direct_vendor, plan,"", from_date, to_date, "", omax_gross_wt]
#                                                     ws.append(l1)
#                                                     print(l1, "222222222222222222")
#                                                     l1 = [front_partcode + "_" + bop , "net_wt", direct_vendor, plan,"", from_date, to_date, "", omax_net_wt]
#                                                     ws.append(l1)
#                                                     l1 = [front_partcode + "_" + bop, "process_cost", direct_vendor, plan,"", from_date, to_date, "", omax_process_cost]
#                                                     ws.append(l1)
#                                                     l1 = [front_partcode + "_" + bop, "dep_cost", direct_vendor, plan,"", from_date, to_date, "", omax_dep_cost]
#                                                     ws.append(l1)
#                                                     l1 = [front_partcode + "_" + bop, "tooling_cost", direct_vendor, plan,"", from_date, to_date, "", omax_tooling_cost]
#                                                     ws.append(l1)
#                                                     l1 = [front_partcode + "_" + bop, "overhead", direct_vendor, plan,"", from_date, to_date, "", omax_overhead]
#                                                     ws.append(l1)
#                                                     l1 = [front_partcode + "_" + bop, "profit_NRMC", direct_vendor, plan,"", from_date, to_date, "", omax_profit_NRMC]
#                                                     ws.append(l1)
#                                                     l1 = [front_partcode + "_" + bop, "profit_process", direct_vendor, plan,"", from_date, to_date, "", omax_profit_process]
#                                                     ws.append(l1)
#                                                     l1 = [front_partcode + "_" + bop, "rej_NRMC", direct_vendor, plan,"", from_date, to_date, "", omax_rej_NRMC]
#                                                     ws.append(l1)
#                                                     l1 = [front_partcode + "_" + bop, "rej_process", direct_vendor, plan,"", from_date, to_date, "", omax_rej_process]
#                                                     ws.append(l1)
#                                                     l1 = [front_partcode + "_" + bop, "fr_NRMC", direct_vendor, plan,"", from_date, to_date, "", omax_fr_NRMC]
#                                                     ws.append(l1)
#                                                     l1 = [front_partcode + "_" + bop, "fr_process", direct_vendor, plan,"", from_date, to_date, "", omax_fr_process]
#                                                     ws.append(l1)

                    
#                     elif "K17" in summary_bop_working_rev_f:
#                         # print(summary_bop_working_rev_f)
#                         val = re.sub("I\$\d+|[^0-9]", "", summary_bop_working_rev_f)
#                         # print(val)
#                         summary_bop_rev_pr = summary_f["K" + str(val)].value
#                         # print(summary_bop_rev)
#                         summary_part_col = summary_f["G" + str(val)].value
#                         # print(summary_part_col)


#                         # for 'OMAX 125cc'
#                         for o in range(33, 760):
#                             omax_part = omax["C" + str(o)].value
#                             omax_price = omax["AI" + str(o)].value
#                             omax_price_f = omax_f["AI" + str(o)].value
#                             # print(omax_price_f)

#                             if omax_part == summary_part_col and summary_bop_rev_pr == omax_price or omax_part == summary_part_col and summary_bop_working_rev == omax_price:
#                                 # print(omax_part)
#                                 # print(omax_price_f)

#                                 omax_price_f = str(omax_price_f)
#                                 omax_price_f = omax_price_f.replace("AI", "").replace("=", "").replace("+", "")
#                                 omax_price_f = omax_f["AI" + str(omax_price_f)].value
#                                 omax_price_f1 = omax_price_f.replace("AI", "").replace("=", "").split("+")
#                                 omax_price_f1.sort(reverse = False)
#                                 # print(omax_price_f1)

#                                 for x in omax_price_f1:
#                                     # print(x)
#                                     omax_price_fx = omax_f["AI" + str(x)].value
#                                     # print(omax_price_fx ,"AI" + str(x))

#                                     if ":" in omax_price_fx:
#                                         omax_price_fx_s = omax_price_fx.replace("SUM", "").replace("=", "").replace("AI", "").replace(")", "").replace("(", "")
#                                         # print(omax_price_fx_s)

#                                         x, y = omax_price_fx_s.split(':')
#                                         # print(x,y)

#                                         for b in range(int(x) - 1, int(x)):
#                                             omax_child = omax["C" + str(b)].value
#                                             if omax_child == None:
#                                                 continue
#                                             elif omax_child == "Part Number":
#                                                 for d in range(int(x) - 2, int(x) - 1):
#                                                     omax_child = omax["C" + str(d)].value
#                                             # print(omax_child, b)

#                                         for c in range(int(x), int(y) + 1):
#                                             omax_sub_child = omax["C" + str(c)].value
#                                             if omax_sub_child == None:
#                                                 omax_sub_child = omax["B" + str(c)].value
#                                             # print(omax_sub_child)

#                                             omax_nrmc = omax["R" + str(c)].value
#                                             omax_gross_wt = omax["E" + str(c)].value
#                                             if omax_gross_wt == None:
#                                                 omax_gross_wt = 0 
#                                             omax_net_wt = omax["F" + str(c)].value
#                                             if omax_net_wt == None:
#                                                 omax_net_wt = 0 
#                                             omax_scrap_wt = omax["G" + str(c)].value
#                                             if omax_scrap_wt == None:
#                                                 omax_scrap_wt = 0 
#                                             omax_rm_grade = omax["H" + str(c)].value
#                                             if omax_rm_grade == None:
#                                                 omax_rm_grade = 0 
#                                             omax_process_cost = omax["T" + str(c)].value + omax["AK" + str(c)].value
#                                             # print(omax["T" + str(c)].value)
#                                             # print(omax["AK" + str(c)].value)
#                                             if omax["U" + str(c)].value == None:
#                                                 omax["U" + str(c)].value = 0
#                                             if omax["AL" + str(c)].value == None:
#                                                 omax["AL" + str(c)].value = 0
#                                             omax_dep_cost = omax["U" + str(c)].value + omax["AL" + str(c)].value
#                                             omax_tooling_cost = omax["V" + str(c)].value
#                                             if omax_tooling_cost == None:
#                                                 omax_tooling_cost = 0 
#                                             omax_overhead = omax["W" + str( c)].value
#                                             if omax_overhead == None:
#                                                 omax_overhead = 0 
#                                             omax_profit_NRMC = omax["X" + str(c)].value
#                                             if omax_profit_NRMC == None:
#                                                 omax_profit_NRMC = 0 
#                                             omax_profit_process = omax["Y" + str(c)].value
#                                             if omax_profit_process == None:
#                                                 omax_profit_process = 0 
#                                             omax_rej_NRMC = omax["Z" + str(c)].value
#                                             if omax_rej_NRMC == None:
#                                                 omax_rej_NRMC = 0 
#                                             omax_rej_process = omax["AA" + str(c)].value
#                                             if omax_rej_process == None:
#                                                 omax_rej_process = 0 
#                                             omax_fr_NRMC = omax["AB" + str(c)].value
#                                             if omax_fr_NRMC == None:
#                                                 omax_fr_NRMC = 0 
#                                             omax_fr_process = omax["AC" + str(c)].value
#                                             if omax_fr_process == None:
#                                                 omax_fr_process = 0 

#                                             omax_no_off = omax["D" + str(c)].value
#                                             omax_bop_cost = omax["M" + str(c)].value

#                                             if omax_gross_wt == None and omax_net_wt == None and omax_process_cost == None and omax_dep_cost == None and omax_tooling_cost == None and omax_overhead == None and omax_profit_NRMC == None and omax_profit_process == None and omax_rej_NRMC == None and omax_rej_process == None and omax_fr_NRMC == None and omax_fr_process == None:
#                                                 continue
#                                             else:
#                                                 l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child) + "_" + str(omax_rm_grade), "gross_wt", direct_vendor, plan,"", from_date, to_date, "", omax_gross_wt]
#                                                 ws.append(l1)
#                                                 l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child) + "_" + str(omax_rm_grade), "net_wt", direct_vendor, plan,"", from_date, to_date, "", omax_net_wt]
#                                                 ws.append(l1)
#                                                 l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child), "process_cost", direct_vendor, plan,"", from_date, to_date, "", omax_process_cost]
#                                                 ws.append(l1)
#                                                 l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child), "dep_cost", direct_vendor, plan,"", from_date, to_date, "", omax_dep_cost]
#                                                 ws.append(l1)
#                                                 l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child), "tooling_cost", direct_vendor, plan,"", from_date, to_date, "", omax_tooling_cost]
#                                                 ws.append(l1)
#                                                 l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child), "overhead", direct_vendor, plan,"", from_date, to_date, "", omax_overhead]
#                                                 ws.append(l1)
#                                                 l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child), "profit_NRMC", direct_vendor, plan,"", from_date, to_date, "", omax_profit_NRMC]
#                                                 ws.append(l1)
#                                                 l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child), "profit_process", direct_vendor, plan,"", from_date, to_date, "", omax_profit_process]
#                                                 ws.append(l1)
#                                                 l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child), "rej_NRMC", direct_vendor, plan,"", from_date, to_date, "", omax_rej_NRMC]
#                                                 ws.append(l1)
#                                                 l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child), "rej_process", direct_vendor, plan,"", from_date, to_date, "", omax_rej_process]
#                                                 ws.append(l1)
#                                                 l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child), "fr_NRMC", direct_vendor, plan,"", from_date, to_date, "", omax_fr_NRMC]
#                                                 ws.append(l1)
#                                                 l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child), "fr_process", direct_vendor, plan,"", from_date, to_date, "", omax_fr_process]
#                                                 ws.append(l1)
                                                
#                                                 if omax_gross_wt == None and omax_net_wt == None:
#                                                     l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child), "NO_OFF", direct_vendor, plan,"", from_date, to_date, "", omax_no_off]
#                                                     ws.append(l1)
#                                                     l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child), "BOP_COST", direct_vendor, plan,"", from_date, to_date, "", omax_bop_cost]
#                                                     ws.append(l1)

#                                     elif "+" in omax_price_fx:
#                                         if omax_price_fx[:2] == "=R":
#                                             # print(omax_price_fx, "ncjdnjcvd")
#                                             omax_price_fx_r = omax_price_fx.replace("R", "").replace("=", "").replace("AK", "").replace("AL", "").replace("AE", "").split("+")
#                                             r = omax_price_fx_r[0]
#                                             # print(r)
#                                             # print(omax_price_fx_r)

#                                             omax_gross_wt = omax["E" + str(r)].value
#                                             if omax_gross_wt == None:
#                                                     omax_gross_wt = 0 
#                                                 # print(omax_gross_wt,"E" + str(r))
#                                             omax_net_wt = omax["F" + str(r)].value
#                                             if omax_net_wt == None:
#                                                 omax_net_wt = 0 
#                                             omax_scrap_wt = omax["G" + str(r)].value
#                                             if omax_scrap_wt == None:
#                                                 omax_scrap_wt = 0 
#                                             omax_rm_grade = omax["H" + str(r)].value
#                                             if omax_rm_grade == None:
#                                                 omax_rm_grade = 0 
#                                             omax_revised_NRMC = omax["R" + str(r)].value
#                                             if omax_revised_NRMC == None:
#                                                 omax_revised_NRMC = 0 
#                                             if omax["AK" + str(r)].value == None : 
#                                                 omax["AK" + str(r)].value = 0
#                                             if omax["T" + str(r)].value == None:
#                                                 omax["T" + str(r)].value = 0
#                                             omax_process_cost = omax["T" + str(r)].value + omax["AK" + str(r)].value
#                                                 # print(omax_process_cost)
                                            
#                                             # print(omax["T" + str(r)].value)
#                                             # print(omax["AK" + str(r)].value)
#                                             if omax["U" + str(r)].value == None: 
#                                                 omax["U" + str(r)].value = 0
#                                             if omax["AL" + str(r)].value == None:
#                                                 omax["AL" + str(r)].value = 0
#                                             omax_dep_cost = omax["U" + str(r)].value + omax["AL" + str(r)].value
#                                             # print(omax_dep_cost)
#                                             omax_tooling_cost = omax["V" + str(r)].value
#                                             if omax_tooling_cost == None:
#                                                 omax_tooling_cost = 0 
#                                             omax_overhead = omax["W" + str(r)].value
#                                             if omax_overhead == None:
#                                                 omax_overhead = 0 
#                                             omax_profit_NRMC = omax["X" + str(r)].value
#                                             if omax_profit_NRMC == None:
#                                                 omax_profit_NRMC = 0 
#                                             omax_profit_process = omax["Y" + str(r)].value
#                                             if omax_profit_process == None:
#                                                 omax_profit_process = 0 
#                                             omax_rej_NRMC = omax["Z" + str(r)].value
#                                             if omax_rej_NRMC == None:
#                                                 omax_rej_NRMC = 0 
#                                             omax_rej_process = omax["AA" + str(r)].value
#                                             if omax_rej_process == None:
#                                                 omax_rej_process = 0 
#                                             omax_fr_NRMC = omax["AB" + str(r)].value
#                                             if omax_fr_NRMC == None:
#                                                 omax_fr_NRMC = 0 
#                                             omax_fr_process = omax["AC" + str(r)].value
#                                             if omax_fr_process == None:
#                                                 omax_fr_process = 0 
#                                             omax_revised_total_cost = omax["AI" + str(r)].value
#                                             if omax_revised_total_cost == None:
#                                                 omax_revised_total_cost = 0 
#                                             omax_rm_grade = str(omax["H" + str(r)].value)
#                                             if omax_rm_grade == None:
#                                                 omax_rm_grade = " "
#                                             # print(omax_rm_grade)
#                                             # print(omax_revised_total_cost, x)
#                                             omax_NRMC = omax["R" + str(r)].value
#                                             if omax_NRMC == None:
#                                                 omax_NRMC = 0 
#                                             # print(omax_NRMC, r)
#                                             bop = omax["C" + str(r)].value
#                                             if bop == None:
#                                                 bop = omax["B" + str(r)].value
#                                                 if bop == "BOP welding":
#                                                     bop = "WELD-01"
#                                                 if bop == "Gauging Cost":
#                                                     bop = "GAUG-01"
#                                             # print(bop, r)

#                                             if omax_gross_wt == None and omax_net_wt == None and omax_process_cost == None and omax_dep_cost == None and omax_tooling_cost == None and omax_overhead == None and omax_profit_NRMC == None and omax_profit_process == None and omax_rej_NRMC == None and omax_rej_process == None and omax_fr_NRMC == None and omax_fr_process == None:
#                                                 continue
#                                             else:
#                                                 l1 = [front_partcode + "_" + bop, "gross_wt", direct_vendor, plan,"", from_date, to_date, "", omax_gross_wt]
#                                                 ws.append(l1)
#                                                 print(l1, "222222222222222222")
#                                                 l1 = [front_partcode + "_" + bop , "net_wt", direct_vendor, plan,"", from_date, to_date, "", omax_net_wt]
#                                                 ws.append(l1)
#                                                 l1 = [front_partcode + "_" + bop, "process_cost", direct_vendor, plan,"", from_date, to_date, "", omax_process_cost]
#                                                 ws.append(l1)
#                                                 l1 = [front_partcode + "_" + bop, "dep_cost", direct_vendor, plan,"", from_date, to_date, "", omax_dep_cost]
#                                                 ws.append(l1)
#                                                 l1 = [front_partcode + "_" + bop, "tooling_cost", direct_vendor, plan,"", from_date, to_date, "", omax_tooling_cost]
#                                                 ws.append(l1)
#                                                 l1 = [front_partcode + "_" + bop, "overhead", direct_vendor, plan,"", from_date, to_date, "", omax_overhead]
#                                                 ws.append(l1)
#                                                 l1 = [front_partcode + "_" + bop, "profit_NRMC", direct_vendor, plan,"", from_date, to_date, "", omax_profit_NRMC]
#                                                 ws.append(l1)
#                                                 l1 = [front_partcode + "_" + bop, "profit_process", direct_vendor, plan,"", from_date, to_date, "", omax_profit_process]
#                                                 ws.append(l1)
#                                                 l1 = [front_partcode + "_" + bop, "rej_NRMC", direct_vendor, plan,"", from_date, to_date, "", omax_rej_NRMC]
#                                                 ws.append(l1)
#                                                 l1 = [front_partcode + "_" + bop, "rej_process", direct_vendor, plan,"", from_date, to_date, "", omax_rej_process]
#                                                 ws.append(l1)
#                                                 l1 = [front_partcode + "_" + bop, "fr_NRMC", direct_vendor, plan,"", from_date, to_date, "", omax_fr_NRMC]
#                                                 ws.append(l1)
#                                                 l1 = [front_partcode + "_" + bop, "fr_process", direct_vendor, plan,"", from_date, to_date, "", omax_fr_process]
#                                                 ws.append(l1)

#                     else:
#                         # print(summary_bop_working_rev_f)
#                         val = re.sub("I\$\d+|[^0-9]", "", summary_bop_working_rev_f)
#                         # print(val)
#                         summary_bop_rev_pr = summary_f["K" + str(val)].value
#                         # print(summary_bop_rev)
#                         summary_part_col = summary_f["G" + str(val)].value
#                         # print(summary_part_col)

#                         # for 'OMAX 125cc'
#                         for o in range(33, 760):
#                             omax_part = omax["C" + str(o)].value
#                             omax_price = omax["AI" + str(o)].value
#                             omax_price_f = omax_f["AI" + str(o)].value
#                             # print(omax_price_f)

#                             if omax_part == summary_part_col and summary_col == omax_price or omax_part == summary_part_col and summary_bop_working_rev == omax_price:
#                                 # print(omax_part)
#                                 # print(omax_price_f)

#                                 omax_price_f = omax_price_f.replace("AI", "").replace("=", "").replace("+", "")
#                                 omax_price_f = omax_f["AI" + str(omax_price_f)].value
#                                 omax_price_f1 = omax_price_f.replace("AI", "").replace("=", "").split("+")
#                                 omax_price_f1.sort(reverse = False)
#                                 # print(omax_price_f1)

#                                 for x in omax_price_f1:
#                                     # print(x)
#                                     omax_price_fx = omax_f["AI" + str(x)].value
#                                     # print(omax_price_fx ,"AI" + str(x))

#                                     if ":" in omax_price_fx:
#                                         omax_price_fx_s = omax_price_fx.replace("SUM", "").replace("=", "").replace("AI", "").replace(")", "").replace("(", "")
#                                         # print(omax_price_fx_s)

#                                         x, y = omax_price_fx_s.split(':')
#                                         # print(x,y)

#                                         for b in range(int(x) - 1, int(x)):
#                                             omax_child = omax["C" + str(b)].value
#                                             if omax_child == None:
#                                                 continue
#                                             elif omax_child == "Part Number":
#                                                 for d in range(int(x) - 2, int(x) - 1):
#                                                     omax_child = omax["C" + str(d)].value
#                                             # print(omax_child, b)

#                                         for c in range(int(x), int(y) + 1):
#                                             omax_sub_child = omax["C" + str(c)].value
#                                             if omax_sub_child == None:
#                                                 omax_sub_child = omax["B" + str(c)].value
#                                             # print(omax_sub_child)

#                                             omax_nrmc = omax["R" + str(c)].value
#                                             omax_gross_wt = omax["E" + str(c)].value
#                                             if omax_gross_wt == None:
#                                                 omax_gross_wt = 0 
#                                             omax_net_wt = omax["F" + str(c)].value
#                                             if omax_net_wt == None:
#                                                 omax_net_wt = 0 
#                                             omax_scrap_wt = omax["G" + str(c)].value
#                                             if omax_scrap_wt == None:
#                                                 omax_scrap_wt = 0 
#                                             omax_rm_grade = omax["H" + str(c)].value
#                                             if omax_rm_grade == None:
#                                                 omax_rm_grade = 0 
#                                             omax_process_cost = omax["T" + str(c)].value + omax["AK" + str(c)].value
#                                             # print(omax["T" + str(c)].value)
#                                             # print(omax["AK" + str(c)].value)
#                                             if omax["U" + str(c)].value == None:
#                                                 omax["U" + str(c)].value = 0
#                                             if omax["AL" + str(c)].value == None:
#                                                 omax["AL" + str(c)].value = 0
#                                             omax_dep_cost = omax["U" + str(c)].value + omax["AL" + str(c)].value
#                                             omax_tooling_cost = omax["V" + str(c)].value
#                                             if omax_tooling_cost == None:
#                                                 omax_tooling_cost = 0 
#                                             omax_overhead = omax["W" + str( c)].value
#                                             if omax_overhead == None:
#                                                 omax_overhead = 0 
#                                             omax_profit_NRMC = omax["X" + str(c)].value
#                                             if omax_profit_NRMC == None:
#                                                 omax_profit_NRMC = 0 
#                                             omax_profit_process = omax["Y" + str(c)].value
#                                             if omax_profit_process == None:
#                                                 omax_profit_process = 0 
#                                             omax_rej_NRMC = omax["Z" + str(c)].value
#                                             if omax_rej_NRMC == None:
#                                                 omax_rej_NRMC = 0 
#                                             omax_rej_process = omax["AA" + str(c)].value
#                                             if omax_rej_process == None:
#                                                 omax_rej_process = 0 
#                                             omax_fr_NRMC = omax["AB" + str(c)].value
#                                             if omax_fr_NRMC == None:
#                                                 omax_fr_NRMC = 0 
#                                             omax_fr_process = omax["AC" + str(c)].value
#                                             if omax_fr_process == None:
#                                                 omax_fr_process = 0 

#                                             omax_no_off = omax["D" + str(c)].value
#                                             omax_bop_cost = omax["M" + str(c)].value

#                                             if omax_gross_wt == None and omax_net_wt == None and omax_process_cost == None and omax_dep_cost == None and omax_tooling_cost == None and omax_overhead == None and omax_profit_NRMC == None and omax_profit_process == None and omax_rej_NRMC == None and omax_rej_process == None and omax_fr_NRMC == None and omax_fr_process == None:
#                                                 continue
#                                             else:
#                                                 l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child) + "_" + str(omax_rm_grade), "gross_wt", direct_vendor, plan,"", from_date, to_date, "", omax_gross_wt]
#                                                 ws.append(l1)
#                                                 l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child) + "_" + str(omax_rm_grade), "net_wt", direct_vendor, plan,"", from_date, to_date, "", omax_net_wt]
#                                                 ws.append(l1)
#                                                 l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child), "process_cost", direct_vendor, plan,"", from_date, to_date, "", omax_process_cost]
#                                                 ws.append(l1)
#                                                 l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child), "dep_cost", direct_vendor, plan,"", from_date, to_date, "", omax_dep_cost]
#                                                 ws.append(l1)
#                                                 l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child), "tooling_cost", direct_vendor, plan,"", from_date, to_date, "", omax_tooling_cost]
#                                                 ws.append(l1)
#                                                 l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child), "overhead", direct_vendor, plan,"", from_date, to_date, "", omax_overhead]
#                                                 ws.append(l1)
#                                                 l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child), "profit_NRMC", direct_vendor, plan,"", from_date, to_date, "", omax_profit_NRMC]
#                                                 ws.append(l1)
#                                                 l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child), "profit_process", direct_vendor, plan,"", from_date, to_date, "", omax_profit_process]
#                                                 ws.append(l1)
#                                                 l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child), "rej_NRMC", direct_vendor, plan,"", from_date, to_date, "", omax_rej_NRMC]
#                                                 ws.append(l1)
#                                                 l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child), "rej_process", direct_vendor, plan,"", from_date, to_date, "", omax_rej_process]
#                                                 ws.append(l1)
#                                                 l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child), "fr_NRMC", direct_vendor, plan,"", from_date, to_date, "", omax_fr_NRMC]
#                                                 ws.append(l1)
#                                                 l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child), "fr_process", direct_vendor, plan,"", from_date, to_date, "", omax_fr_process]
#                                                 ws.append(l1)
                                                
#                                                 if omax_gross_wt == None and omax_net_wt == None:
#                                                     l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child), "NO_OFF", direct_vendor, plan,"", from_date, to_date, "", omax_no_off]
#                                                     ws.append(l1)
#                                                     l1 = [str(front_partcode) + "_" + str(omax_child) + "_" + str(omax_sub_child), "BOP_COST", direct_vendor, plan,"", from_date, to_date, "", omax_bop_cost]
#                                                     ws.append(l1)

#                                     elif "+" in omax_price_fx:
#                                         if omax_price_fx[:2] == "=R":
#                                             # print(omax_price_fx, "ncjdnjcvd")
#                                             omax_price_fx_r = omax_price_fx.replace("R", "").replace("=", "").replace("AK", "").replace("AL", "").replace("AE", "").split("+")
#                                             r = omax_price_fx_r[0]
#                                             # print(r)
#                                             # print(omax_price_fx_r)

#                                             omax_gross_wt = omax["E" + str(r)].value
#                                             if omax_gross_wt == None:
#                                                     omax_gross_wt = 0 
#                                                 # print(omax_gross_wt,"E" + str(r))
#                                             omax_net_wt = omax["F" + str(r)].value
#                                             if omax_net_wt == None:
#                                                 omax_net_wt = 0 
#                                             omax_scrap_wt = omax["G" + str(r)].value
#                                             if omax_scrap_wt == None:
#                                                 omax_scrap_wt = 0 
#                                             omax_rm_grade = omax["H" + str(r)].value
#                                             if omax_rm_grade == None:
#                                                 omax_rm_grade = 0 
#                                             omax_revised_NRMC = omax["R" + str(r)].value
#                                             if omax_revised_NRMC == None:
#                                                 omax_revised_NRMC = 0 
#                                             if omax["AK" + str(r)].value == None : 
#                                                 omax["AK" + str(r)].value = 0
#                                             if omax["T" + str(r)].value == None:
#                                                 omax["T" + str(r)].value = 0
#                                             omax_process_cost = omax["T" + str(r)].value + omax["AK" + str(r)].value
#                                                 # print(omax_process_cost)
                                            
#                                             # print(omax["T" + str(r)].value)
#                                             # print(omax["AK" + str(r)].value)
#                                             if omax["U" + str(r)].value == None: 
#                                                 omax["U" + str(r)].value = 0
#                                             if omax["AL" + str(r)].value == None:
#                                                 omax["AL" + str(r)].value = 0
#                                             omax_dep_cost = omax["U" + str(r)].value + omax["AL" + str(r)].value
#                                             # print(omax_dep_cost)
#                                             omax_tooling_cost = omax["V" + str(r)].value
#                                             if omax_tooling_cost == None:
#                                                 omax_tooling_cost = 0 
#                                             omax_overhead = omax["W" + str(r)].value
#                                             if omax_overhead == None:
#                                                 omax_overhead = 0 
#                                             omax_profit_NRMC = omax["X" + str(r)].value
#                                             if omax_profit_NRMC == None:
#                                                 omax_profit_NRMC = 0 
#                                             omax_profit_process = omax["Y" + str(r)].value
#                                             if omax_profit_process == None:
#                                                 omax_profit_process = 0 
#                                             omax_rej_NRMC = omax["Z" + str(r)].value
#                                             if omax_rej_NRMC == None:
#                                                 omax_rej_NRMC = 0 
#                                             omax_rej_process = omax["AA" + str(r)].value
#                                             if omax_rej_process == None:
#                                                 omax_rej_process = 0 
#                                             omax_fr_NRMC = omax["AB" + str(r)].value
#                                             if omax_fr_NRMC == None:
#                                                 omax_fr_NRMC = 0 
#                                             omax_fr_process = omax["AC" + str(r)].value
#                                             if omax_fr_process == None:
#                                                 omax_fr_process = 0 
#                                             omax_revised_total_cost = omax["AI" + str(r)].value
#                                             if omax_revised_total_cost == None:
#                                                 omax_revised_total_cost = 0 
#                                             omax_rm_grade = omax["H" + str(r)].value
#                                             if omax_rm_grade == None:
#                                                 omax_rm_grade = " "
#                                             # print(omax_rm_grade)
#                                             # print(omax_revised_total_cost, x)
#                                             omax_NRMC = omax["R" + str(r)].value
#                                             if omax_NRMC == None:
#                                                 omax_NRMC = 0 
#                                             # print(omax_NRMC, r)
#                                             bop = omax["C" + str(r)].value
#                                             if bop == None:
#                                                 bop = omax["B" + str(r)].value
#                                                 if bop == "BOP welding":
#                                                     bop = "WELD-01"
#                                                 if bop == "Gauging Cost":
#                                                     bop = "GAUG-01"
#                                             # print(bop, r)

#                                             if omax_gross_wt == None and omax_net_wt == None and omax_process_cost == None and omax_dep_cost == None and omax_tooling_cost == None and omax_overhead == None and omax_profit_NRMC == None and omax_profit_process == None and omax_rej_NRMC == None and omax_rej_process == None and omax_fr_NRMC == None and omax_fr_process == None:
#                                                 continue
#                                             else:
#                                                 l1 = [front_partcode + "_" + bop, "gross_wt", direct_vendor, plan,"", from_date, to_date, "", omax_gross_wt]
#                                                 ws.append(l1)
#                                                 print(l1, "222222222222222222")
#                                                 l1 = [front_partcode + "_" + bop , "net_wt", direct_vendor, plan,"", from_date, to_date, "", omax_net_wt]
#                                                 ws.append(l1)
#                                                 l1 = [front_partcode + "_" + bop, "process_cost", direct_vendor, plan,"", from_date, to_date, "", omax_process_cost]
#                                                 ws.append(l1)
#                                                 l1 = [front_partcode + "_" + bop, "dep_cost", direct_vendor, plan,"", from_date, to_date, "", omax_dep_cost]
#                                                 ws.append(l1)
#                                                 l1 = [front_partcode + "_" + bop, "tooling_cost", direct_vendor, plan,"", from_date, to_date, "", omax_tooling_cost]
#                                                 ws.append(l1)
#                                                 l1 = [front_partcode + "_" + bop, "overhead", direct_vendor, plan,"", from_date, to_date, "", omax_overhead]
#                                                 ws.append(l1)
#                                                 l1 = [front_partcode + "_" + bop, "profit_NRMC", direct_vendor, plan,"", from_date, to_date, "", omax_profit_NRMC]
#                                                 ws.append(l1)
#                                                 l1 = [front_partcode + "_" + bop, "profit_process", direct_vendor, plan,"", from_date, to_date, "", omax_profit_process]
#                                                 ws.append(l1)
#                                                 l1 = [front_partcode + "_" + bop, "rej_NRMC", direct_vendor, plan,"", from_date, to_date, "", omax_rej_NRMC]
#                                                 ws.append(l1)
#                                                 l1 = [front_partcode + "_" + bop, "rej_process", direct_vendor, plan,"", from_date, to_date, "", omax_rej_process]
#                                                 ws.append(l1)
#                                                 l1 = [front_partcode + "_" + bop, "fr_NRMC", direct_vendor, plan,"", from_date, to_date, "", omax_fr_NRMC]
#                                                 ws.append(l1)
#                                                 l1 = [front_partcode + "_" + bop, "fr_process", direct_vendor, plan,"", from_date, to_date, "", omax_fr_process]
#                                                 ws.append(l1)

                    



        # Summary-Sub assy-Child Part
        # for s in range(11, 36):
#             sub_part = sub["D" + str(s)].value
#             sub_price = sub["AU" + str(s)].value
#             if front_partcode == sub_part and revised_price == sub_price:
#                 # print("sub_part", sub_part)
#                 # print("sub_price", sub_price)

#                 sub_interest_cost = sub["X" + str(s)].value
#                 if sub_interest_cost == None:
#                     sub_interest_cost = 0
#                 sub_depriciation_cost = sub["Y" + str(s)].value
#                 if sub_depriciation_cost == None:
#                     sub_depriciation_cost = 0
#                 sub_Wire_and_Co2_gas = sub["Z" + str(s)].value
#                 if sub_Wire_and_Co2_gas == None:
#                     sub_Wire_and_Co2_gas = 0
#                 sub_other_cost = sub["AA" + str(s)].value
#                 if sub_other_cost == None:
#                     sub_other_cost = 0
#                 sub_welding_fix_cost = sub["AB" + str(s)].value
#                 if sub_welding_fix_cost == None:
#                     sub_welding_fix_cost = 0
#                 sub_press_shop_cost = sub["AC" + str(s)].value
#                 if sub_press_shop_cost == None:
#                     sub_press_shop_cost = 0
#                 sub_OH_at_30_cost = sub["AD" + str(s)].value
#                 sub_profit_cost = sub["AE" + str(s)].value
#                 sub_rejection_at_rm_and_conv_cost = sub["AF" + str(s)].value
#                 sub_paking_exp_cost = sub["AG" + str(s)].value
#                 sub_freight_cost = sub["AH" + str(s)].value
#                 sub_ammortization_cost = sub["AI" + str(s)].value
#                 if sub_ammortization_cost == None:
#                      sub_ammortization_cost = 0
#                 sub_total = sub["AJ" + str(s)].value
#                 if sub_total == None:
#                      sub_total = 0                  
#                 sub_additional_price = sub["BL" + str(s)].value
#                 if sub_additional_price == None:
#                     sub_additional_price = 0
#                 sub_rm_sec_operaton_charges = sub["AK" + str(s)].value
#                 sub_delta_due_to_labour_on_welding = sub["AL" + str(s)].value
#                 sub_delta_due_to_labour_on_press_shop = sub["AM" + str(s)].value
#                 sub_delta_due_to_labour_on_other_ops = sub["AN" + str(s)].value
#                 sub_delta_due_to_interest_on_press_shop = sub["AO" + str(s)].value
#                 if sub_delta_due_to_interest_on_press_shop == None:
#                      sub_delta_due_to_interest_on_press_shop = 0
#                 sub_welding_cost_revised = sub["K" + str(s)].value
#                 if sub_welding_cost_revised == None:
#                     sub_welding_cost_revised = 0
#                 sub_labour = sub["O" + str(s)].value
#                 if sub_labour == None:
#                     sub_labour = 0
#                 sub_power = sub["S" + str(s)].value
#                 if sub_power == None:
#                     sub_power = 0
#                 sub_welding_cost_revised = sub["H" + str(s)].value
#                 if sub_welding_cost_revised == None:
#                     sub_welding_cost_revised = 0

#                 # AAWD-HNBL Child part Working 
#                 for a in range(29, 209):
#                     aawd_part = aawd["D" + str(a)].value
#                     aawd_price = aawd["AL" + str(a)].value

#                     if aawd_part == sub_part and aawd_price == sub_welding_cost_revised:
#                         aawd_price_f = aawd_f["AL" + str(a)].value
#                         # print(aawd_price_f)
#                         if ":" in aawd_price_f:
#                             aawd_price_f_s = aawd_price_f.replace("SUM", "").replace("=", "").replace("AL", "").replace(")", "").replace("(", "")
#                             # print(aawd_price_f_s)
#                             x, y = aawd_price_f_s.split(':')
#                             # print(x,y)

#                             for b in range(int(x) - 1, int(x)):
#                                 aawd_child = aawd["B" + str(b)].value
#                                 # print(aawd_child, b)
                            
#                             for c in range(int(x), int(y) + 1):
#                                 aawd_sub_child = aawd["C" + str(c)].value
#                                 # print(aawd_sub_child)

#                                 # aawd_process_cost_conversion_inc = aawd["AN" + str(a)].value
#                                 # # print(aawd_process_cost_conversion_inc)
#                                 # aawd_int_rationalisation = aawd["AO" + str(a)].value

                                
#                                 aawd_gross_wt = aawd["H" + str(c)].value
#                                 if aawd_gross_wt == None:
#                                     aawd_gross_wt = 0 
#                                 aawd_net_wt = aawd["I" + str(c)].value
#                                 if aawd_net_wt == None:
#                                     aawd_net_wt = 0 
#                                 aawd_scrap_wt = aawd["J" + str(c)].value
#                                 if aawd_scrap_wt == None:
#                                     aawd_scrap_wt = 0 
#                                 aawd_rm_grade = aawd["K" + str(c)].value
#                                 if aawd_rm_grade == None:
#                                     aawd_rm_grade = 0 

#                                 if aawd["W" + str(c)].value == None:
#                                     aawd["W" + str(c)].value = 0
#                                 if aawd["AN" + str(c)].value == None:
#                                     aawd["AN" + str(c)].value = 0
#                                 aawd_process_cost = aawd["W" + str(c)].value + aawd["AN" + str(c)].value
#                                 if aawd["X" + str(c)].value == None:
#                                     aawd["X" + str(c)].value = 0
#                                 if aawd["AO" + str(c)].value == None:
#                                     aawd["AO" + str(c)].value = 0
#                                 aawd_dep_cost = aawd["X" + str(c)].value + aawd["AO" + str(c)].value
#                                 aawd_tooling_cost = aawd["Y" + str(c)].value
#                                 if aawd_tooling_cost == None:
#                                     aawd_tooling_cost = 0 
#                                 aawd_overhead = aawd["Z" + str( c)].value
#                                 if aawd_overhead == None:
#                                     aawd_overhead = 0 
#                                 aawd_profit_NRMC = aawd["AA" + str(c)].value
#                                 if aawd_profit_NRMC == None:
#                                     aawd_profit_NRMC = 0 
#                                 aawd_profit_process = aawd["AB" + str(c)].value
#                                 if aawd_profit_process == None:
#                                     aawd_profit_process = 0 
#                                 aawd_rej_NRMC = aawd["AC" + str(c)].value
#                                 if aawd_rej_NRMC == None:
#                                     aawd_rej_NRMC = 0 
#                                 aawd_rej_process = aawd["AD" + str(c)].value
#                                 if aawd_rej_process == None:
#                                     aawd_rej_process = 0 
#                                 aawd_fr_NRMC = aawd["AE" + str(c)].value
#                                 if aawd_fr_NRMC == None:
#                                     aawd_fr_NRMC = 0 
#                                 aawd_fr_process = aawd["AF" + str(c)].value
#                                 if aawd_fr_process == None:
#                                     aawd_fr_process = 0 

#                                 aawd_no_off = aawd["G" + str(c)].value
#                                 aawd_bop_cost = aawd["R" + str(c)].value

#                                 if aawd_gross_wt == None and aawd_net_wt == None and aawd_process_cost == None and aawd_dep_cost == None and aawd_tooling_cost == None and aawd_overhead == None and aawd_profit_NRMC == None and aawd_profit_process == None and aawd_rej_NRMC == None and aawd_rej_process == None and aawd_fr_NRMC == None and aawd_fr_process == None:
#                                     continue
#                                 else: 
#                                     l1 = [str(front_partcode) + "_" + str(aawd_child) + "_" + str(aawd_sub_child) + "_" + str(aawd_rm_grade), "gross_wt", direct_vendor, plan,"", from_date, to_date, "", aawd_gross_wt]
#                                     ws.append(l1)
#                                     l1 = [str(front_partcode) + "_" + str(aawd_child) + "_" + str(aawd_sub_child) + "_" + str(aawd_rm_grade), "net_wt", direct_vendor, plan,"", from_date, to_date, "", aawd_net_wt]
#                                     ws.append(l1)
#                                     l1 = [str(front_partcode) + "_" + str(aawd_child) + "_" + str(aawd_sub_child), "process_cost", direct_vendor, plan,"", from_date, to_date, "", aawd_process_cost]
#                                     ws.append(l1)
#                                     l1 = [str(front_partcode) + "_" + str(aawd_child) + "_" + str(aawd_sub_child), "dep_cost", direct_vendor, plan,"", from_date, to_date, "", aawd_dep_cost]
#                                     ws.append(l1)
#                                     l1 = [str(front_partcode) + "_" + str(aawd_child) + "_" + str(aawd_sub_child), "tooling_cost", direct_vendor, plan,"", from_date, to_date, "", aawd_tooling_cost]
#                                     ws.append(l1)
#                                     l1 = [str(front_partcode) + "_" + str(aawd_child) + "_" + str(aawd_sub_child), "overhead", direct_vendor, plan,"", from_date, to_date, "", aawd_overhead]
#                                     ws.append(l1)
#                                     l1 = [str(front_partcode) + "_" + str(aawd_child) + "_" + str(aawd_sub_child), "profit_NRMC", direct_vendor, plan,"", from_date, to_date, "", aawd_profit_NRMC]
#                                     ws.append(l1)
#                                     l1 = [str(front_partcode) + "_" + str(aawd_child) + "_" + str(aawd_sub_child), "profit_process", direct_vendor, plan,"", from_date, to_date, "", aawd_profit_process]
#                                     ws.append(l1)
#                                     l1 = [str(front_partcode) + "_" + str(aawd_child) + "_" + str(aawd_sub_child), "rej_NRMC", direct_vendor, plan,"", from_date, to_date, "", aawd_rej_NRMC]
#                                     ws.append(l1)
#                                     l1 = [str(front_partcode) + "_" + str(aawd_child) + "_" + str(aawd_sub_child), "rej_process", direct_vendor, plan,"", from_date, to_date, "", aawd_rej_process]
#                                     ws.append(l1)
#                                     l1 = [str(front_partcode) + "_" + str(aawd_child) + "_" + str(aawd_sub_child), "fr_NRMC", direct_vendor, plan,"", from_date, to_date, "", aawd_fr_NRMC]
#                                     ws.append(l1)
#                                     l1 = [str(front_partcode) + "_" + str(aawd_child) + "_" + str(aawd_sub_child), "fr_process", direct_vendor, plan,"", from_date, to_date, "", aawd_fr_process]
#                                     ws.append(l1)
                                    
#                                     if aawd_gross_wt == None and aawd_net_wt == None:
#                                         l1 = [str(front_partcode) + "_" + str(aawd_child) + "_" + str(aawd_sub_child), "NO_OFF", direct_vendor, plan,"", from_date, to_date, "", aawd_no_off]
#                                         ws.append(l1)
#                                         l1 = [str(front_partcode) + "_" + str(aawd_child) + "_" + str(aawd_sub_child), "BOP_COST", direct_vendor, plan,"", from_date, to_date, "", aawd_bop_cost]
#                                         ws.append(l1)




# wb.save(r"F:\sequelstring\Excel\NEELMETAL_OUTPUT\Output\sub_assy_master.xlsx")





