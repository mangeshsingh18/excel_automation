import re
import openpyxl
import psycopg2
from openpyxl import Workbook, load_workbook

#  Main Sheet
wb = Workbook()
# ws = wb.create_chartsheet
ws = openpyxl.load_workbook(r"F:\sequelstring\Excel\NEELMETAL_OUTPUT\Output\frame_hierachy.xlsx", data_only=True)
ws = wb['Sheet']

# FRT SHEET 01102021
wb1 = openpyxl.load_workbook(r"F:\sequelstring\Excel\NEELMETAL_OUTPUT\PIR_FRAME_NCR_NEELMETAL_01.01.2022.xlsx", data_only=True)
front_sheet = wb1['FRT SHEET 01102021']

# SUMMARY-FRAMEBODY
wb2 = openpyxl.load_workbook(r"F:\sequelstring\Excel\NEELMETAL_OUTPUT\PIR_FRAME_NCR_NEELMETAL_01.01.2022.xlsx", data_only=True)
summary_t = wb2['SUMMARY-FRAMEBODY']

wb3 = openpyxl.load_workbook(r"F:\sequelstring\Excel\NEELMETAL_OUTPUT\PIR_FRAME_NCR_NEELMETAL_01.01.2022.xlsx", data_only=False)
summary_f = wb3['SUMMARY-FRAMEBODY']

# ANNEX-VTV
wb4 = openpyxl.load_workbook(r"F:\sequelstring\Excel\NEELMETAL_OUTPUT\PIR_FRAME_NCR_NEELMETAL_01.01.2022.xlsx", data_only=True)
annex = wb4['ANNEX-VTV']

# OMAX 125cc
wb5 = openpyxl.load_workbook(r"F:\sequelstring\Excel\NEELMETAL_OUTPUT\PIR_FRAME_NCR_NEELMETAL_01.01.2022.xlsx", data_only=True)
omax_t = wb5['OMAX 125cc']

wb6 = openpyxl.load_workbook(r"F:\sequelstring\Excel\NEELMETAL_OUTPUT\PIR_FRAME_NCR_NEELMETAL_01.01.2022.xlsx", data_only=False)
omax_f = wb6['OMAX 125cc']

headers = ['bom_hierarchy', 'direct_sup', 'purchase_group', 'plan',	'from_date', 'to_date',	'partcode', 'partcode_type1_oe',
'partcode_level2','partcode_type2_rm_bop_vtv_inm',	'partcode_level3', 'partcode_type3_rm_bop_vtv_inm',	'partcode_level4', 'partcode_type4_rm_bop_vtv_inm',	'partcode_level5','partcode_type5_rm_bop_vtv_inm','partcode_level6','partcode_type6_rm_bop_vtv_inm_base_reference_pc']
ws.append(headers)



H_column = annex["H" + str(4)].value
I_column = annex["I" + str(4)].value
J_column = annex["J" + str(4)].value
K_column = annex["K" + str(4)].value
L_column = annex["L" + str(4)].value



# for front sheet
for i in range(5, 93):
    category = front_sheet["S" + str(i)].value
    # print(category)

    if category == "Frame Body ":
        front_partcode = front_sheet["G" + str(i)].value
        revised_price = front_sheet["P" + str(i)].value
        direct_vendor = front_sheet["F" + str(i)].value
        plan = front_sheet["E" + str(i)].value
        from_date = "01.01.22"
        to_date = "31.12.9999"
        # print(front_partcode, i)
        # print(revised_price)

        l1 = [ "", direct_vendor,"", plan, from_date, to_date, front_partcode, "OE"]
        ws.append(l1)

        # for summary framebody
        for j in range(9, 73):
            summary_part = summary_t["G" + str(j)].value
            summary_price1 = summary_t["BN" + str(j)].value
            summary_price2 = summary_t["AT" + str(j)].value
            summary_price1_f = summary_f["BN" + str(j)].value
            summary_price2_f = summary_f["AT" + str(j)].value
            # print(summary_part)
            # print(summary_price2_f)

            if front_partcode == summary_part and revised_price == summary_price1 or summary_price2 == revised_price:
                # print("front_partcode", front_partcode)
                # print("summary_part", summary_part)
                # print("revised_price", revised_price)
                # print("summary_price", summary_price)

                
                summary_vtv_f = summary_f["N" + str(j)].value
                # print(summary_vtv_f)
                # summary_vtv = summary_t["N" + str(j)].value
                # print(front_partcode)
                # print(type(summary_vtv))

                if "=" in str(summary_vtv_f):
                    # print(summary_vtv_f)
                    # summary_vtv_f = re.findall("N\d+", summary_vtv_f)
                    # print(summary_vtv_f)
                    # for x in summary_vtv_f:
                    #     summary_vtv_f = summary_f[x].value
                    #     # print(bop_set_f)

                    if 'ANNEX-VTV' in summary_vtv_f:
                        summ_vtv = summary_t["G" + str(j)].value
                        # print(summ_vtv,1111)

                    else:
                        if "$" in summary_vtv_f:
                            # print(summary_vtv_f)

                            val = re.sub('[^0-9]', "", summary_vtv_f)
                            # print(val)
                            summ_vtv = summary_t["G" + str(val)].value
                            # print(summ_vtv,2222)

                    

                    if summ_vtv == H_column: 
                        v1 = "H"
                    elif summ_vtv == I_column: 
                        v1 = "I"
                    elif summ_vtv == J_column: 
                        v1 = "J"
                    elif summ_vtv == K_column: 
                        v1 = "K"
                    elif summ_vtv == L_column: 
                        v1 = "L"

                    for a in range(6, 17):
                        no = annex[v1 + str(a)].value 
                        # print(1)
                        if no != None:
                            # print(1)
                            ann_partno = annex["C" + str(a)].value
                            # print(ann_partno)
                            po_no = annex[v1 + str(a)].value

                            l1 = [ "", direct_vendor,"", plan, from_date, to_date, front_partcode, "OE", ann_partno, "VTV"]
                            ws.append(l1)
                            print(l1)


                summary_additional_price = summary_t["BL" + str(j)].value
                if summary_additional_price == None:
                    summary_additional_price = 0
                summary_rm_sec_operaton_charges = summary_t["AM" + str(20)].value
                summary_interest_cost = summary_t["Z" + str(j)].value
                summary_depriciation_cost = summary_t["AA" + str(j)].value
                summary_Wire_and_Co2_gas = summary_t["AB" + str(j)].value
                summary_other_cost = summary_t["AC" + str(j)].value
                summary_welding_fix_cost = summary_t["AD" + str(j)].value
                if summary_welding_fix_cost == None:
                    summary_welding_fix_cost = 0
                summary_press_shop_cost = summary_t["AE" + str(j)].value
                summary_OH_at_30_cost = summary_t["AF" + str(j)].value
                summary_profit_cost = summary_t["AG" + str(j)].value
                summary_rejection_at_rm_and_conv_cost = summary_t["AH" + str(j)].value
                summary_paking_exp_cost = summary_t["AI" + str(j)].value
                summary_freight_cost = summary_t["AJ" + str(j)].value
                summary_ammortization_cost = summary_t["AK" + str(j)].value
                summary_power_welding = summary_t["V" + str(j)].value
                summary_power_press_shop = summary_t["W" + str(j)].value
                summary_labour_welding = summary_t["R" + str(j)].value
                summary_labour_press_shop = summary_t["S" + str(j)].value
                summary_labour_other_assy_operation = summary_t["T" + str(j)].value
                # print(summary_additional_price)

                summary_bop_revised_f = summary_f["K" + str(j)].value
                summary_bop_revised = summary_t["K" + str(j)].value
                summary_partcode = summary_t["G" + str(j)].value

                # print("summary_partcode", summary_partcode)
                # print("summary_bop_revised", summary_bop_revised)

                # for 'OMAX 125cc'
                for o in range(53, 760):
                    omax_partcode = omax_t["C" + str(o)].value
                    omax_price = omax_t["AI" + str(o)].value

                    if omax_partcode == summary_partcode and omax_price == summary_bop_revised:
                        omax_partcode_f = omax_f["C" + str(o)].value
                        # print("omax_partcode_f", omax_partcode_f)
                        omax_price_f = omax_f["AI" + str(o)].value
                        # print("omax_price_f", omax_price_f)

                        omax_price_f = omax_price_f.replace("AI", "").replace("=", "")
                        omax_price_f = omax_f["AI" + str(omax_price_f)].value
                        # print(omax_price_f)
                        omax_price_f1 = omax_price_f.replace("AI", "").replace("=", "").split("+")
                        omax_price_f1.sort(reverse = False)
                        # print(omax_price_f1)

                        for x in omax_price_f1:
                            # print(x)
                            omax_price_fx = omax_f["AI" + str(x)].value
                            # print("omax_price_fx", omax_price_fx)

                            if ":" in omax_price_fx:
                                omax_price_fx_s = omax_price_fx.replace("SUM", "").replace("=", "").replace("AI", "").replace(")", "").replace("(", "")
                                # print(omax_price_fx_s)

                                x, y = omax_price_fx_s.split(':')
                                # print(x,y)

                                for b in range(int(x) - 1, int(x)):
                                    omax_child = omax_t["C" + str(b)].value
                                    if omax_child == None:
                                        continue
                                    elif omax_child == "Part Number":
                                        for d in range(int(x) - 2, int(x) - 1):
                                            omax_child = omax_t["C" + str(d)].value
                                    # print(omax_child, b)

                                for c in range(int(x), int(y) + 1):
                                    omax_sub_child = omax_t["C" + str(c)].value
                                    if omax_sub_child == None:
                                        omax_sub_child = omax_t["B" + str(c)].value
                                    # print(omax_sub_child)

                                    

                                    omax_gross_wt = omax_t["E" + str(c)].value
                                    if omax_gross_wt == None:
                                        omax_gross_wt = 0 
                                    omax_net_wt = omax_t["F" + str(c)].value
                                    if omax_net_wt == None:
                                        omax_net_wt = 0 
                                    omax_scrap_wt = omax_t["G" + str(c)].value
                                    if omax_scrap_wt == None:
                                        omax_scrap_wt = 0 
                                    omax_rm_grade = omax_t["H" + str(c)].value
                                    if omax_rm_grade == None:
                                        omax_rm_grade = 0 

                                    if omax_t["T" + str(c)].value == None:
                                        omax_t["T" + str(c)].value = 0
                                    if omax_t["AK" + str(c)].value == None:
                                        omax_t["AK" + str(c)].value = 0
                                    omax_process_cost = omax_t["T" + str(c)].value + omax_t["AK" + str(c)].value
                                    if omax_t["U" + str(c)].value == None:
                                        omax_t["U" + str(c)].value = 0
                                    if omax_t["AL" + str(c)].value == None:
                                        omax_t["AL" + str(c)].value = 0
                                    omax_dep_cost = omax_t["U" + str(c)].value + omax_t["AL" + str(c)].value
                                    omax_tooling_cost = omax_t["V" + str(c)].value
                                    if omax_tooling_cost == None:
                                        omax_tooling_cost = 0 
                                    omax_overhead = omax_t["W" + str( c)].value
                                    if omax_overhead == None:
                                        omax_overhead = 0 
                                    omax_profit_NRMC = omax_t["X" + str(c)].value
                                    if omax_profit_NRMC == None:
                                        omax_profit_NRMC = 0 
                                    omax_profit_process = omax_t["Y" + str(c)].value
                                    if omax_profit_process == None:
                                        omax_profit_process = 0 
                                    omax_rej_NRMC = omax_t["Z" + str(c)].value
                                    if omax_rej_NRMC == None:
                                        omax_rej_NRMC = 0 
                                    omax_rej_process = omax_t["AA" + str(c)].value
                                    if omax_rej_process == None:
                                        omax_rej_process = 0 
                                    omax_fr_NRMC = omax_t["AB" + str(c)].value
                                    if omax_fr_NRMC == None:
                                        omax_fr_NRMC = 0 
                                    omax_fr_process = omax_t["AC" + str(c)].value
                                    if omax_fr_process == None:
                                        omax_fr_process = 0 
                                    omax_RM_grade = omax_t["H" + str(c)].value
                                    if omax_RM_grade == None:
                                        omax_RM_grade = " "
                                    omax_nrmc = omax_t["R" + str(c)].value
                                    omax_no_off = omax_t["D" + str(c)].value
                                    omax_bop_cost = omax_t["M" + str(c)].value

                                                                       
                                    # for hierachy
                                    if omax_gross_wt == None and omax_net_wt == None and omax_process_cost == None and omax_dep_cost == None and omax_tooling_cost == None and omax_overhead == None and omax_profit_NRMC == None and omax_profit_process == None and omax_rej_NRMC == None and omax_rej_process == None and omax_fr_NRMC == None and omax_fr_process == None:
                                        continue
                                    else:
                                        l1 = ["", direct_vendor, "", plan, from_date, to_date, front_partcode, "OE", omax_child, "INM"]
                                        ws.append(l1)
                                        l1 = ["", direct_vendor, "", plan, from_date, to_date, front_partcode, "OE", omax_child, "INM", omax_sub_child, "INM"]
                                        ws.append(l1)
                                        l1 = ["", direct_vendor, "", plan, from_date, to_date, front_partcode, "OE", omax_child, "INM", omax_sub_child, "INM", omax_RM_grade, "RM"]
                                        ws.append(l1)
                                        l1 = ["", direct_vendor, "", plan, from_date, to_date, front_partcode, "OE", omax_child, "INM", omax_sub_child, "INM", "Scrap", "RM"]
                                        ws.append(l1)

                                        if omax_gross_wt == None and omax_net_wt == None:
                                            l1 = ["", direct_vendor, "", plan, from_date, to_date, front_partcode, "OE", omax_child, "INM"]
                                            ws.append(l1)
                                            l1 = ["" , direct_vendor, "", plan, from_date, to_date, front_partcode, "OE", omax_child, "INM", omax_sub_child, "BOP"]
                                            ws.append(l1)
                                            print(l1)


                            elif "+" in str(omax_price_fx):
                                omax_price_fx = omax_price_fx.replace("R", "").replace("=", "").replace("AK", "").replace("AL", "").replace("AE", "").split("+")
                                # print(omax_price_fx)
                                r = omax_price_fx[0]
                                # print(r)  

                                omax_gross_wt = omax_t["E" + str(r)].value
                                omax_net_wt = omax_t["F" + str(r)].value
                                
                                if omax_t["T" + str(r)].value == None:
                                    omax_t["T" + str(r)].value = 0
                                if omax_t["AK" + str(r)].value == None:
                                    omax_t["AK" + str(r)].value = 0
                                omax_process_cost = omax_t["T" + str(r)].value + omax_t["AK" + str(r)].value
                                if omax_t["U" + str(r)].value == None:
                                    omax_t["U" + str(r)].value = 0
                                if omax_t["AL" + str(r)].value == None:
                                    omax_t["AL" + str(r)].value = 0
                                omax_dep_cost = omax_t["U" + str(r)].value + omax_t["AL" + str(r)].value
                                omax_tooling_cost = omax_t["V" + str(r)].value
                                omax_overhead = omax_t["W" + str( r)].value
                                omax_profit_NRMC = omax_t["X" + str(r)].value
                                omax_profit_process = omax_t["Y" + str(r)].value
                                omax_rej_NRMC = omax_t["Z" + str(r)].value
                                omax_rej_process = omax_t["AA" + str(r)].value
                                omax_fr_NRMC = omax_t["AB" + str(r)].value
                                omax_fr_process = omax_t["AC" + str(r)].value

                                bop = omax_t["C" + str(r)].value
                                if bop == None:
                                    bop = omax_t["B" + str(r)].value
                                    if bop == "BOP welding":
                                        bop = "WELD-01"
                                    if bop == "Gauging Cost":
                                        bop = "GAUG-01"
                                # print(bop, r)

                                l1 = ["" , direct_vendor, "", plan, from_date, to_date, front_partcode, "OE",bop, "INM"]
                                ws.append(l1)

                ws.append({})
                                                    
                                    


            # # print("________________DONE__________________")




















wb.save(r"F:\sequelstring\Excel\NEELMETAL_OUTPUT\Output\frame_hierachy.xlsx")