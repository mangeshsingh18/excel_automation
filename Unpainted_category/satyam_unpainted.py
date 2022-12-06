import re
import openpyxl 
from openpyxl import load_workbook, Workbook

wb = Workbook()
# ws = wb.create_chartsheet

# Master Sheet
ws = openpyxl.load_workbook(r"F:\sequelstring\Excel\TASK 3\unpainted\satyam_pir_unpainted_master.xlsx", data_only=True)

# Heirachy Sheet
# ws = openpyxl.load_workbook(r"F:\sequelstring\Excel\TASK 3\unpainted\satyam_pir_unpainted_heirachy.xlsx", data_only=True)
ws = wb['Sheet']

wb1 = load_workbook(r"F:\sequelstring\Excel\TASK 3\PIR_FRAME_HALOL_SATYAM_01.01.2022 - Copy.xlsx", data_only=True)
front = wb1['Frt Sheet -Satyam']

wb2 = load_workbook(r"F:\sequelstring\Excel\TASK 3\PIR_FRAME_HALOL_SATYAM_01.01.2022 - Copy.xlsx", data_only=False)
front_f = wb2['Frt Sheet -Satyam']

wb3 = load_workbook(r"F:\sequelstring\Excel\TASK 3\PIR_FRAME_HALOL_SATYAM_01.01.2022 - Copy.xlsx", data_only=True)
painted = wb3['Painted PIR']

wb4 = load_workbook(r"F:\sequelstring\Excel\TASK 3\PIR_FRAME_HALOL_SATYAM_01.01.2022 - Copy.xlsx", data_only=False)
painted_f = wb4['Painted PIR']

wb5 = load_workbook(r"F:\sequelstring\Excel\TASK 3\PIR_FRAME_HALOL_SATYAM_01.01.2022 - Copy.xlsx", data_only=True)
summary = wb5['SUMMARY-FRAMEBODY SATYAM (2)']

wb6 = load_workbook(r"F:\sequelstring\Excel\TASK 3\PIR_FRAME_HALOL_SATYAM_01.01.2022 - Copy.xlsx", data_only=False)
summary_f = wb6['SUMMARY-FRAMEBODY SATYAM (2)']

wb7 = load_workbook(r"F:\sequelstring\Excel\TASK 3\PIR_FRAME_HALOL_SATYAM_01.01.2022 - Copy.xlsx", data_only=True)
satyam = wb7['SATYAM']

wb8 = load_workbook(r"F:\sequelstring\Excel\TASK 3\PIR_FRAME_HALOL_SATYAM_01.01.2022 - Copy.xlsx", data_only=False)
satyam_f = wb8['SATYAM']

wb9 = openpyxl.load_workbook(r"F:\sequelstring\Excel\TASK 3\PIR_FRAME_HALOL_SATYAM_01.01.2022.xlsx", data_only=True)
annex = wb9['ANNEX-3 VTV ']

# Master sheet
headers = ['bom_hierarchy',	'master', 'direct_sup',	'plan',	'frequency', 'from_date',	'to_date',	'purchase_group',	'value',	'percentage',	'input_currency',	'output_currency',
        'unit',	'from_city',	'to_city',	'from_period',	'to_period',	'forward_exchange',	'leap_master',	'indicator',	'rm_exclude_flag',	'osp_conversion',	'osp_freight']

ws.append(headers)

# # Heirachy Sheet
# headers = ['bom_hierarchy', 'direct_sup', 'purchase_group', 'plan',	'from_date', 'to_date',	'partcode', 'partcode_type1_oe',
#     'partcode_level2','partcode_type2_rm_bop_vtv_inm',	'partcode_level3', 'partcode_type3_rm_bop_vtv_inm',	'partcode_level4', 'partcode_type4_rm_bop_vtv_inm',	'partcode_level5','partcode_type5_rm_bop_vtv_inm','partcode_level6','partcode_type6_rm_bop_vtv_inm_base_reference_pc']

# ws.append(headers)



# for front sheet
for i in range(4, 64):
    category = front["N" + str(i)].value
    # print(category)
    if category == "UNPAINTED":
        partcode = front["G" + str(i)].value
        # print(partcode, i)
        price = front["Q" + str(i)].value
        # print(price)
        price_f = front_f["Q" + str(i)].value
        # print(price_f)
        direct_vendor = front["F" + str(i)].value
        plan = front["E" + str(i)].value
        from_date = "01.01.22"
        to_date = "31.12.9999"

        # for summary
        for j in range(9, 21):
            summ_part = summary["D" + str(j)].value
            summ_price = summary["AY" + str(j)].value
             
            # for annex
            if summ_part == partcode and summ_price == price:

                ws.append({})

                # hierachy sheet
                # l1 = [ "", direct_vendor,"", plan, from_date, to_date, partcode, "OE"]
                # ws.append(l1)

                summ_price_f = summary_f["AY" + str(j)].value
                # print(summ_price_f)

                bop_revised = summary["H" + str(j)].value 
                bop_revised_f = summary_f["H" + str(j)].value 
                # print(bop_revised_f) 

                summ_vtv = summary_f["K" + str(j)].value
                # print(summ_vtv)
                
                if 'ANNEX' in summ_vtv:
                    summ_col = summary["D" + str(j)].value
                    # print(summ_col)
                elif 'ANNEX' not in summ_vtv:
                    # print(summ_vtv)
                    val = re.sub("[^0-9]", "",summ_vtv)
                    summ_col = summary["D" + str(val)].value
                    # print(summ_col)
                
                for i in ["K", "L", "M", "N", "O", "P", "Q"]:
                    vtv_part = annex[i + str(4)].value
                    # print(vtv_part)

                    if summ_col == vtv_part:
                        # print(vtv_part)

                        for a in range(6, 18):
                            no = annex[i + str(a)].value
                            if no != None:
                                ann_partno = annex["C" + str(a)].value
                                # print(ann_partno)
                                po_no = annex[str(i) + str(a)].value
                                # print(po_no)

                                # Master Sheet
                                list1 = [partcode + "_" + ann_partno, "no_off", direct_vendor, plan,"", from_date, to_date, "", po_no]
                                ws.append(list1)
                                print(list1)

                                # # Heirachy Sheet
                                # l1 = [ "", direct_vendor,"", plan, from_date, to_date, partcode, "OE", ann_partno, "VTV"]
                                # ws.append(l1)
                                # print(l1)


                for_HF_deluxe = summary["AS" + str(j)].value
                cost_impact_due_to_rpt = summary["AT" + str(j)].value
                if cost_impact_due_to_rpt == None:
                    cost_impact_due_to_rpt = 0
                
                HM5V_conversion = summary["AO" + str(j)].value
                RM_secondary_operation_charges = summary["AM" + str(j)].value
                delta_due_to_increase_in_weld_wire = summary["AI" + str(j)].value
                power_welding = summary["S" + str(j)].value
                power_press_shop = summary["T" + str(j)].value
                labour_welding = summary["O" + str(j)].value
                labour_press_shop = summary["P" + str(j)].value
                other_assy = summary["Q" + str(j)].value
                interest_cost = summary["W" + str(j)].value
                depreciation_cost = summary["X" + str(j)].value
                Wire_and_Co2_gas = summary["Y" + str(j)].value
                other_cost = summary["Z" + str(j)].value
                welding_fixtures = summary["AA" + str(j)].value
                press_shop = summary["AB" + str(j)].value
                OH_at_30 = summary["AC" + str(j)].value
                profit = summary["AD" + str(j)].value
                rejection_on_RM_and_conv = summary["AE" + str(j)].value
                packing_exp = summary["AF" + str(j)].value
                freight_existing = summary["AG" + str(j)].value

                # Master Sheet
                list1 = [partcode, "interest_cost", direct_vendor, plan,"", from_date, to_date, "", interest_cost]
                ws.append(list1)
                list1 = [partcode, "depreciation_cost", direct_vendor, plan,"", from_date, to_date, "", depreciation_cost]
                ws.append(list1)
                list1 = [partcode, "Wire_and_Co2_gas", direct_vendor, plan,"", from_date, to_date, "", Wire_and_Co2_gas]
                ws.append(list1)
                list1 = [partcode, "other_cost", direct_vendor, plan,"", from_date, to_date, "", other_cost]
                ws.append(list1)
                list1 = [partcode, "welding_fixtures", direct_vendor, plan,"", from_date, to_date, "", welding_fixtures]
                ws.append(list1)
                list1 = [partcode, "press_shop", direct_vendor, plan,"", from_date, to_date, "", power_press_shop]
                ws.append(list1)
                list1 = [partcode, "OH_at_30", direct_vendor, plan,"", from_date, to_date, "", OH_at_30]
                ws.append(list1)
                list1 = [partcode, "profit", direct_vendor, plan,"", from_date, to_date, "", profit]
                ws.append(list1)
                list1 = [partcode, "rejection_on_RM_and_conv", direct_vendor, plan,"", from_date, to_date, "", rejection_on_RM_and_conv]
                ws.append(list1)
                list1 = [partcode, "packing_exp", direct_vendor, plan,"", from_date, to_date, "", packing_exp]
                ws.append(list1)
                list1 = [partcode, "fright_existing", direct_vendor, plan,"", from_date, to_date, "", freight_existing]
                ws.append(list1)
                list1 = [partcode, "delta_due_to_increase_in_weld_wire", direct_vendor, plan,"", from_date, to_date, "", delta_due_to_increase_in_weld_wire]
                ws.append(list1)
                list1 = [partcode, "power_welding", direct_vendor, plan,"", from_date, to_date, "", power_welding]
                ws.append(list1)
                list1 = [partcode, "power_press_shop", direct_vendor, plan,"", from_date, to_date, "", power_press_shop]
                ws.append(list1)
                list1 = [partcode, "labour_welding", direct_vendor, plan,"", from_date, to_date, "", labour_welding]
                ws.append(list1)
                list1 = [partcode, "labour_press_shop", direct_vendor, plan,"", from_date, to_date, "", labour_press_shop]
                ws.append(list1)
                list1 = [partcode, "other_assy", direct_vendor, plan,"", from_date, to_date, "", other_assy]
                ws.append(list1)
                list1 = [partcode, "cost_impact_due_to_RPT", direct_vendor, plan,"", from_date, to_date, "", cost_impact_due_to_rpt]
                ws.append(list1)
                list1 = [partcode, "RM_secondary_operation_charges", direct_vendor, plan,"", from_date, to_date, "", RM_secondary_operation_charges]
                ws.append(list1)
                list1 = [partcode, "2%_for_HF_deluxe", direct_vendor, plan,"", from_date, to_date, "", for_HF_deluxe]
                ws.append(list1)
                list1 = [partcode, "1.5%_HM5V_conversion", direct_vendor, plan,"", from_date, to_date, "", HM5V_conversion]
                ws.append(list1)
                

                # satyam
                for o in range(31, 1523):
                    satyam_partno = satyam["C" + str(o)].value
                    satyam_revised_price = satyam["AI" + str(o)].value
                    satyam_revised_price_f = satyam_f["AI" + str(o)].value

                    if summ_part == satyam_partno and bop_revised == satyam_revised_price:
                        # print(satyam_partno, "@@@@@@@@@@@@@@@@@@@@@@@", o)
                        satyam_revised_price_f = re.sub("[^0-9]", "",satyam_revised_price_f)
                        satyam_rev_price_f = satyam_f["AI" + str(satyam_revised_price_f)].value
                        satyam_rev_price_f = re.sub("[A-Z]+|=", "", satyam_rev_price_f)
                        satyam_rev_price_f1 = satyam_rev_price_f.split("+")
                        satyam_rev_price_f1.sort(reverse = False)
                        # print(satyam_rev_price_f1)

                        for x in satyam_rev_price_f1:
                            # print(x)
                            revise_total_cost_f = satyam_f["AI" + str(x)].value
                            # print(revise_total_cost_f)

                            if ":" in revise_total_cost_f:
                                revise_total_cost_f = re.sub("[A-Z]+|=|\(|\)", "", revise_total_cost_f)
                                # print(revise_total_cost_f)
                                x, y = revise_total_cost_f.split(':')
                                # print(x,y)

                                for b in range(int(x) - 1, int(x)):
                                    satyam_child = satyam["C" + str(b)].value
                                    if satyam_child == None:
                                        continue
                                    elif satyam_child == "Part Number":
                                        for d in range(int(x) - 2, int(x) - 1):
                                            satyam_child = satyam["C" + str(d)].value
                                    # print(satyam_child, b)

                                for c in range(int(x), int(y) + 1):
                                    satyam_sub_child = satyam["C" + str(c)].value
                                    if satyam_sub_child == None:
                                        satyam_sub_child = satyam["B" + str(c)].value
                                    # print(satyam_sub_child, c)


                                    satyam_gross_wt = satyam["E" + str(c)].value
                                    if satyam_gross_wt == None:
                                        satyam_gross_wt = 0 
                                    satyam_net_wt = satyam["F" + str(c)].value
                                    if satyam_net_wt == None:
                                        satyam_net_wt = 0 
                                    if satyam["T" + str(c)].value == None:
                                        satyam["T" + str(c)].value = 0
                                    if satyam["AK" + str(c)].value == None:
                                        satyam["AK" + str(c)].value = 0
                                    satyam_process_cost = satyam["T" + str(c)].value + satyam["AK" + str(c)].value
                                    if satyam["U" + str(c)].value == None:
                                        satyam["U" + str(c)].value = 0
                                    if satyam["AM" + str(c)].value == None:
                                        satyam["AM" + str(c)].value = 0
                                    satyam_dep_cost = satyam["U" + str(c)].value + satyam["AM" + str(c)].value
                                    satyam_tooling_cost = satyam["V" + str(c)].value
                                    if satyam_tooling_cost == None:
                                        satyam_tooling_cost = 0 
                                    satyam_overhead = satyam["W" + str( c)].value
                                    if satyam_overhead == None:
                                        satyam_overhead = 0 
                                    satyam_profit_NRMC = satyam["X" + str(c)].value
                                    if satyam_profit_NRMC == None:
                                        satyam_profit_NRMC = 0 
                                    satyam_profit_process = satyam["Y" + str(c)].value
                                    if satyam_profit_process == None:
                                        satyam_profit_process = 0 
                                    satyam_rej_NRMC = satyam["Z" + str(c)].value
                                    if satyam_rej_NRMC == None:
                                        satyam_rej_NRMC = 0 
                                    satyam_rej_process = satyam["AA" + str(c)].value
                                    if satyam_rej_process == None:
                                        satyam_rej_process = 0 
                                    satyam_fr_NRMC = satyam["AB" + str(c)].value
                                    if satyam_fr_NRMC == None:
                                        satyam_fr_NRMC = 0 
                                    satyam_fr_process = satyam["AC" + str(c)].value
                                    if satyam_fr_process == None:
                                        satyam_fr_process = 0 
                                    satyam_GRMC = satyam["P" + str(c)].value
                                    satyam_scrape_cost = satyam["Q" + str(c)].value
                                    satyam_no_off = satyam["D" + str(c)].value
                                    satyam_NRMC = satyam["R" + str(c)].value
                                    satyam_rm_grade = satyam["H" + str(c)].value
                                    if satyam_rm_grade == None:
                                        satyam_rm_grade = ""
                                    satyam_scrape_wt = satyam["G" + str(c)].value
                                    if satyam_scrape_wt == None:
                                        satyam_scrape_wt = 0
                                    #for bop cost
                                    satyam_rm_rate = satyam["J" + str(c)].value 
                                    if satyam_rm_rate == None:
                                        satyam_rm_rate = 0

                                    # Master Sheet
                                    if satyam_gross_wt == 0 and satyam_net_wt == 0 and satyam_process_cost == 0 and satyam_dep_cost == 0 and satyam_tooling_cost == 0 and satyam_overhead == 0 and satyam_profit_NRMC == 0 and satyam_profit_process == 0 and satyam_rej_NRMC == 0 and satyam_rej_process == 0 and satyam_fr_NRMC == 0 and satyam_fr_process == 0:
                                        continue
                                    else:
                                        if satyam_child != None:
                                            list1 = [str(partcode) + "_" + str(satyam_child) + "_" + str(satyam_sub_child) + "_" + satyam_rm_grade, "gross_wt", direct_vendor, plan,"", from_date, to_date, "", satyam_gross_wt]
                                            ws.append(list1)
                                            print(list1 , "11111111111111111111")

                                            list1 = [str(partcode) + "_" + str(satyam_child) + "_" + str(satyam_sub_child) + "_" + satyam_rm_grade, "net_wt", direct_vendor, plan,"", from_date, to_date, "", satyam_net_wt]
                                            ws.append(list1)
                                            print(list1 , "11111111111111111111")

                                            list1 = [str(partcode) + "_" + str(satyam_child) + "_" + str(satyam_sub_child), "process_cost", direct_vendor, plan,"", from_date, to_date, "", satyam_process_cost]
                                            ws.append(list1)
                                            print(list1 , "11111111111111111111")

                                            list1 = [str(partcode) + "_" + str(satyam_child) + "_" + str(satyam_sub_child), "dep_cost", direct_vendor, plan,"", from_date, to_date, "", satyam_dep_cost]
                                            ws.append(list1)
                                            list1 = [str(partcode) + "_" + str(satyam_child) + "_" + str(satyam_sub_child), "tooling_cost", direct_vendor, plan,"", from_date, to_date, "", satyam_tooling_cost]
                                            ws.append(list1)
                                            list1 = [str(partcode) + "_" + str(satyam_child) + "_" + str(satyam_sub_child), "overhead", direct_vendor, plan,"", from_date, to_date, "", satyam_overhead]
                                            ws.append(list1)
                                            list1 = [str(partcode) + "_" + str(satyam_child) + "_" + str(satyam_sub_child), "profit_NRMC", direct_vendor, plan,"", from_date, to_date, "", satyam_profit_NRMC]
                                            ws.append(list1)
                                            list1 = [str(partcode) + "_" + str(satyam_child) + "_" + str(satyam_sub_child), "profit_process", direct_vendor, plan,"", from_date, to_date, "", satyam_profit_process]
                                            ws.append(list1)
                                            list1 = [str(partcode) + "_" + str(satyam_child) + "_" + str(satyam_sub_child), "rej_NRMC", direct_vendor, plan,"", from_date, to_date, "", satyam_rej_NRMC]
                                            ws.append(list1)
                                            list1 = [str(partcode) + "_" + str(satyam_child) + "_" + str(satyam_sub_child), "rej_process", direct_vendor, plan,"", from_date, to_date, "", satyam_rej_process]
                                            ws.append(list1)
                                            list1 = [str(partcode) + "_" + str(satyam_child) + "_" + str(satyam_sub_child), "fr_NRMC", direct_vendor, plan,"", from_date, to_date, "", satyam_fr_NRMC]
                                            ws.append(list1)
                                            list1 = [str(partcode) + "_" + str(satyam_child) + "_" + str(satyam_sub_child), "fr_process", direct_vendor, plan,"", from_date, to_date, "", satyam_fr_process]
                                            ws.append(list1)
                                        else:
                                            list1 = [str(partcode) + "_" + str(satyam_sub_child) + "_" + satyam_rm_grade, "gross_wt", direct_vendor, plan,"", from_date, to_date, "", satyam_gross_wt]
                                            ws.append(list1)
                                            print(list1 , "11111111111111111111")

                                            list1 = [str(partcode) + "_" + str(satyam_sub_child) + "_" + satyam_rm_grade, "net_wt", direct_vendor, plan,"", from_date, to_date, "", satyam_net_wt]
                                            ws.append(list1)
                                            print(list1 , "11111111111111111111")

                                            list1 = [str(partcode) + "_" + str(satyam_sub_child), "process_cost", direct_vendor, plan,"", from_date, to_date, "", satyam_process_cost]
                                            ws.append(list1)
                                            print(list1 , "11111111111111111111")

                                            list1 = [str(partcode) + "_" + str(satyam_sub_child), "dep_cost", direct_vendor, plan,"", from_date, to_date, "", satyam_dep_cost]
                                            ws.append(list1)
                                            list1 = [str(partcode) + "_" + str(satyam_sub_child), "tooling_cost", direct_vendor, plan,"", from_date, to_date, "", satyam_tooling_cost]
                                            ws.append(list1)
                                            list1 = [str(partcode) + "_" + str(satyam_sub_child), "overhead", direct_vendor, plan,"", from_date, to_date, "", satyam_overhead]
                                            ws.append(list1)
                                            list1 = [str(partcode) + "_" + str(satyam_sub_child), "profit_NRMC", direct_vendor, plan,"", from_date, to_date, "", satyam_profit_NRMC]
                                            ws.append(list1)
                                            list1 = [str(partcode) + "_" + str(satyam_sub_child), "profit_process", direct_vendor, plan,"", from_date, to_date, "", satyam_profit_process]
                                            ws.append(list1)
                                            list1 = [str(partcode) + "_" + str(satyam_sub_child), "rej_NRMC", direct_vendor, plan,"", from_date, to_date, "", satyam_rej_NRMC]
                                            ws.append(list1)
                                            list1 = [str(partcode) + "_" + str(satyam_sub_child), "rej_process", direct_vendor, plan,"", from_date, to_date, "", satyam_rej_process]
                                            ws.append(list1)
                                            list1 = [str(partcode) + "_" + str(satyam_sub_child), "fr_NRMC", direct_vendor, plan,"", from_date, to_date, "", satyam_fr_NRMC]
                                            ws.append(list1)
                                            list1 = [str(partcode) + "_" + str(satyam_sub_child), "fr_process", direct_vendor, plan,"", from_date, to_date, "", satyam_fr_process]
                                            ws.append(list1)
        
                                        if satyam_gross_wt == 0 and satyam_net_wt == 0 and satyam_scrape_cost == None:
                                            l1 = [str(partcode) + "_" + str(satyam_child) + "_" + str(satyam_sub_child), "NO_OFF", direct_vendor, plan,"", from_date, to_date, "", satyam_no_off]
                                            ws.append(l1)
                                            l1 = [str(partcode) + "_" + str(satyam_child) + "_" + str(satyam_sub_child), "BOP_COST", direct_vendor, plan,"", from_date, to_date, "", satyam_rm_rate]
                                            ws.append(l1)

                                    # # Heirachy Sheet
                                    # if satyam_gross_wt == 0 and satyam_net_wt == 0 and satyam_process_cost == 0 and satyam_dep_cost == 0 and satyam_tooling_cost == 0 and satyam_overhead == 0 and satyam_profit_NRMC == 0 and satyam_profit_process == 0 and satyam_rej_NRMC == 0 and satyam_rej_process == 0 and satyam_fr_NRMC == 0 and satyam_fr_process == 0:
                                    #     continue
                                    # else:
                                    #     l1 = ["", direct_vendor, "", plan, from_date, to_date, partcode, "OE", satyam_child, "INM"]
                                    #     ws.append(l1)
                                    #     l1 = ["", direct_vendor, "", plan, from_date, to_date, partcode, "OE", satyam_child, "INM", satyam_sub_child, "INM"]
                                    #     ws.append(l1)
                                    #     l1 = ["", direct_vendor, "", plan, from_date, to_date, partcode, "OE", satyam_child, "INM", satyam_sub_child, "INM", satyam_rm_grade, "RM"]
                                    #     ws.append(l1)
                                    #     l1 = ["", direct_vendor, "", plan, from_date, to_date, partcode, "OE", satyam_child, "INM", satyam_sub_child, "INM", "Scrap", "RM"]
                                    #     ws.append(l1)

                                    #     if satyam_gross_wt == 0 and satyam_net_wt == 0:
                                    #         l1 = ["", direct_vendor, "", plan, from_date, to_date, partcode, "OE", satyam_child, "INM"]
                                    #         ws.append(l1)
                                    #         l1 = ["" , direct_vendor, "", plan, from_date, to_date, partcode, "OE", satyam_child, "INM", satyam_sub_child, "BOP"]
                                    #         ws.append(l1)
                    



                            elif "+" in revise_total_cost_f:
                                # print(revise_total_cost_f) 
                                revise_total_cost_f = re.sub("[A-Z]+|=|\(|\)", "", revise_total_cost_f)  
                                revise_total_cost_f = revise_total_cost_f.split("+")
                                # print(revise_total_cost_f) 
                                r = revise_total_cost_f[0]
                                # print(r)

                                if satyam["T" + str(r)].value == None:
                                    satyam["T" + str(r)].value = 0
                                if satyam["AK" + str(r)].value == None:
                                    satyam["AK" + str(r)].value = 0
                                satyam_process_rost = satyam["T" + str(r)].value + satyam["AK" + str(r)].value
                                if satyam["U" + str(r)].value == None:
                                    satyam["U" + str(r)].value = 0
                                if satyam["AM" + str(r)].value == None:
                                    satyam["AM" + str(r)].value = 0
                                satyam_dep_cost = satyam["U" + str(r)].value + satyam["AM" + str(r)].value
                                satyam_tooling_cost = satyam["V" + str(r)].value
                                if satyam_tooling_cost == None:
                                    satyam_tooling_cost = 0 
                                satyam_overhead = satyam["W" + str( r)].value
                                if satyam_overhead == None:
                                    satyam_overhead = 0 
                                satyam_profit_NRMC = satyam["X" + str(r)].value
                                if satyam_profit_NRMC == None:
                                    satyam_profit_NRMC = 0 
                                satyam_profit_process = satyam["Y" + str(r)].value
                                if satyam_profit_process == None:
                                    satyam_profit_process = 0 
                                satyam_rej_NRMC = satyam["Z" + str(r)].value
                                if satyam_rej_NRMC == None:
                                    satyam_rej_NRMC = 0 
                                satyam_rej_process = satyam["AA" + str(r)].value
                                if satyam_rej_process == None:
                                    satyam_rej_process = 0 
                                satyam_fr_NRMC = satyam["AB" + str(r)].value
                                if satyam_fr_NRMC == None:
                                    satyam_fr_NRMC = 0 
                                satyam_fr_process = satyam["AC" + str(r)].value
                                if satyam_fr_process == None:
                                    satyam_fr_process = 0 
                                satyam_NRMC = satyam["R" + str(r)].value
                                bop = satyam["C" + str(r)].value
                                if bop == None:
                                    bop = satyam["B" + str(r)].value
                                    if bop == "BOP welding":
                                        bop = "WELD-01"
                                    if bop == "Additional BOP welding (2.05 Inch)":
                                        continue
                                    if bop == "Gauging Cost":
                                        bop = "GAUG-01"
                                # print(bop, r)

                                # For Master Sheet
                                if satyam_gross_wt == 0 and satyam_net_wt == 0 and satyam_process_cost == 0 and satyam_dep_cost == 0 and satyam_tooling_cost == 0 and satyam_overhead == 0 and satyam_profit_NRMC == 0 and satyam_profit_process == 0 and satyam_rej_NRMC == 0 and satyam_rej_process == 0 and satyam_fr_NRMC == 0 and satyam_fr_process == 0:
                                    continue
                                else:
                                    list1 = [str(partcode) + "_" + bop, "gross_wt", direct_vendor, plan,"", from_date, to_date, "", satyam_gross_wt]
                                    ws.append(list1)
                                    print(list1, "222222222222222222")
                                    list1 = [partcode + "_" + bop , "net_wt", direct_vendor, plan,"", from_date, to_date, "", satyam_net_wt]
                                    ws.append(list1)
                                    list1 = [partcode + "_" + bop, "process_cost", direct_vendor, plan,"", from_date, to_date, "", satyam_process_cost]
                                    ws.append(list1)
                                    list1 = [partcode + "_" + bop, "dep_cost", direct_vendor, plan,"", from_date, to_date, "", satyam_dep_cost]
                                    ws.append(list1)
                                    list1 = [partcode + "_" + bop, "tooling_cost", direct_vendor, plan,"", from_date, to_date, "", satyam_tooling_cost]
                                    ws.append(list1)
                                    list1 = [partcode + "_" + bop, "overhead", direct_vendor, plan,"", from_date, to_date, "", satyam_overhead]
                                    ws.append(list1)
                                    list1 = [partcode + "_" + bop, "profit_NRMC", direct_vendor, plan,"", from_date, to_date, "", satyam_profit_NRMC]
                                    ws.append(list1)
                                    list1 = [partcode + "_" + bop, "profit_process", direct_vendor, plan,"", from_date, to_date, "", satyam_profit_process]
                                    ws.append(list1)
                                    list1 = [partcode + "_" + bop, "rej_NRMC", direct_vendor, plan,"", from_date, to_date, "", satyam_rej_NRMC]
                                    ws.append(list1)
                                    list1 = [partcode + "_" + bop, "rej_process", direct_vendor, plan,"", from_date, to_date, "", satyam_rej_process]
                                    ws.append(list1)
                                    list1 = [partcode + "_" + bop, "fr_NRMC", direct_vendor, plan,"", from_date, to_date, "", satyam_fr_NRMC]
                                    ws.append(list1)
                                    list1 = [partcode + "_" + bop, "fr_process", direct_vendor, plan,"", from_date, to_date, "", satyam_fr_process]
                                    ws.append(list1)

                                # # For Heirachy Sheet
                                # l1 = ["" , direct_vendor, "", plan, from_date, to_date, partcode, "OE", bop, "INM"]
                                # ws.append(l1)
                                    
                            

# Master Sheet
wb.save(r"F:\sequelstring\Excel\TASK 3\unpainted\satyam_pir_unpainted_master.xlsx")

# Heirachy Sheet
# wb.save(r"F:\sequelstring\Excel\TASK 3\unpainted\satyam_pir_unpainted_heirachy.xlsx")











